import time
import math
import pyvisa
import sys
import numpy as np
import openpyxl
from openpyxl import Workbook
import threading
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import mplcursors
from matplotlib.figure import Figure
from matplotlib.ticker import FuncFormatter


class MeasurementApp:
    def __init__(self):
        """
        Initialize the application:
         - Set up the VISA resource manager and list connected devices.
         - Define VISA addresses for each instrument (ECL laser, wavelength meter, spectrum analyzer, Keithley, etc.).
         - Initialize data containers and threading events.
         - Create the main Tkinter window.
        """
        # Initialize the VISA resource manager and list connected devices (optional)
        self.rm = pyvisa.ResourceManager()
        print("Connected devices:", self.rm.list_resources())

        # Define VISA addresses for the instruments
        self.ecl_adapter_GPIB = 'GPIB0::10::INSTR'         # ECL laser (should be constant)
        self.wavelength_meter_GPIB = 'GPIB0::20::INSTR'      # Wavelength meter
        self.spectrum_analyzer_GPIB = 'GPIB0::18::INSTR'     # Spectrum analyzer
        self.keithley_GPIB = 'GPIB0::24::INSTR'              # Keithley source meter
        self.power_sensor_GPIB = 'GPIB0::13::INSTR'  # Anritsu power sensor
        self.voa_GPIB = 'GPIB0::26::INSTR'                   # VOA

        # Instrument objects (to be opened later)
        self.ecl_adapter = None
        self.wavelength_meter = None
        self.spectrum_analyzer = None
        self.keithley = None
        self.power_sensor = None
        self.voa = None

        # Data containers for measurements and calibration
        self.steps = []
        self.beat_freqs = []
        self.laser_4_wavelengths = []
        self.beat_freq_and_power = []  # List of tuples: (beat_freq, raw RF power, photocurrent, VOA P actual)
        self.calibrated_rf = []
        self.photo_currents = []
        self.rf_loss = []
        self.rf_probe_loss = []
        self.rf_link_loss = []
        self.powers = []
        self.p_actuals = []
        self.looping = False

        # Threading events for controlling data collection and plot updates
        self.stop_event = threading.Event()
        self.data_ready_event = threading.Event()

        # Initialize main Tkinter window (full-screen, or "zoomed")
        self.root = tk.Tk()
        self.root.title("Measurement and Plotting GUI")
        self.root.geometry("1200x800")
        self.root.state('zoomed')

        # Create GUI components and plots
        self.create_gui()
        self.setup_plots()

        # Setup closing protocol and plot updating loop
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.after(100, self.update_plots)

        self.open_instruments()

    def create_gui(self):
        """
        Create the left input frame (with text fields, buttons, file selectors, etc.)
        and the right frame where the Matplotlib plots are embedded.
        Also includes basic type-checked GUI inputs.
        """
        # Create a frame for user inputs on the left
        self.input_frame = ttk.Frame(self.root, width=200)
        self.input_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

        # Create a frame for plots on the right
        self.plot_frame = ttk.Frame(self.root)
        self.plot_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # --- Input fields with type-checked variables ---
        ttk.Label(self.input_frame, text="Starting WL Laser 3 (nm):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.laser_3_var = tk.DoubleVar(value=1550)  # Default value for laser 3 wavelength
        self.laser_3_entry = ttk.Entry(self.input_frame, textvariable=self.laser_3_var)
        self.laser_3_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(self.input_frame, text="Starting WL Laser 4 (nm):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.laser_4_var = tk.DoubleVar(value=1548)  # Default value for laser 4 wavelength
        self.laser_4_entry = ttk.Entry(self.input_frame, textvariable=self.laser_4_var)
        self.laser_4_entry.grid(row=1, column=1, padx=5, pady=5)

        # Checkbox for enabling automatic initial beat frequency search
        ttk.Label(self.input_frame, text="Enable Automatic Initial Beat Frequency Search:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.enable_search_var = tk.BooleanVar(value=True)
        self.enable_search_checkbox = ttk.Checkbutton(self.input_frame, variable=self.enable_search_var)
        self.enable_search_checkbox.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        # Pause after initial search option
        ttk.Label(self.input_frame, text="Pause After Initial Search:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.pause_after_search_var = tk.BooleanVar(value=False)
        self.pause_after_search_cb = ttk.Checkbutton(self.input_frame, variable=self.pause_after_search_var)
        self.pause_after_search_cb.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        # Starting beat frequency input
        ttk.Label(self.input_frame, text="Starting Beat Frequency (GHz):").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.start_freq_var = tk.DoubleVar(value=0)
        self.start_freq_entry = ttk.Entry(self.input_frame, textvariable=self.start_freq_var)
        self.start_freq_entry.grid(row=4, column=1, padx=5, pady=5)

        # Ending beat frequency input
        ttk.Label(self.input_frame, text="Ending Beat Frequency (GHz):").grid(row=5, column=0, padx=5, pady=5, sticky="e")
        self.end_freq_var = tk.DoubleVar(value=0)
        self.end_freq_entry = ttk.Entry(self.input_frame, textvariable=self.end_freq_var)
        self.end_freq_entry.grid(row=5, column=1, padx=5, pady=5)

        # Number of steps in the frequency sweep
        ttk.Label(self.input_frame, text="Number of Steps:").grid(row=6, column=0, padx=5, pady=5, sticky="e")
        self.num_steps_var = tk.IntVar(value=0)
        self.num_steps_entry = ttk.Entry(self.input_frame, textvariable=self.num_steps_var)
        self.num_steps_entry.grid(row=6, column=1, padx=5, pady=5)

        # Delay between steps in the sweep
        ttk.Label(self.input_frame, text="Delay Between Steps (s):").grid(row=7, column=0, padx=5, pady=5, sticky="e")
        self.delay_var = tk.DoubleVar(value=3.5)
        self.delay_entry = ttk.Entry(self.input_frame, textvariable=self.delay_var)
        self.delay_entry.grid(row=7, column=1, padx=5, pady=5)

        
        # Resume button (disabled until pause is selected)
        self.resume_button = ttk.Button(self.input_frame, text="Resume Measurements", command=self.resume_after_pause, state=tk.DISABLED)
        self.resume_button.grid(row=20, column=0, columnspan=2, pady=10)
        ttk.Label(
            self.input_frame,
            text="NOTE: This will resume measurement after the initial beat–freq search if “Pause After Initial Search” is enabled."
        ).grid(row=21, column=0, columnspan=2, pady=2)

        # Toggle resume button when pause checkbox changes
        self.pause_after_search_var.trace_add('write', lambda *_: self.toggle_resume_button())

        # File selector for RF Link Loss (.s2p file) calibration
        ttk.Label(self.input_frame, text="RF Link Loss (.s2p):").grid(row=8, column=0, padx=5, pady=5, sticky="e")
        self.s2p_file_var = tk.StringVar()
        self.s2p_file_entry = ttk.Entry(self.input_frame, textvariable=self.s2p_file_var, width=20)
        self.s2p_file_entry.grid(row=8, column=1, padx=5, pady=5)
        self.s2p_file_button = ttk.Button(self.input_frame, text="Browse", command=self.select_s2p_file)
        self.s2p_file_button.grid(row=8, column=2, padx=5, pady=5)

        # File selector for RF Probe Loss (.xlsx file) calibration
        ttk.Label(self.input_frame, text="RF Probe Loss (.xlsx):").grid(row=9, column=0, padx=5, pady=5, sticky="e")
        self.excel_file_var = tk.StringVar()
        self.excel_file_entry = ttk.Entry(self.input_frame, textvariable=self.excel_file_var, width=20)
        self.excel_file_entry.grid(row=9, column=1, padx=5, pady=5)
        self.excel_file_button = ttk.Button(self.input_frame, text="Browse", command=self.select_excel_file)
        self.excel_file_button.grid(row=9, column=2, padx=5, pady=5)

        # Control buttons: Start, Stop, and Reset
        self.start_button = ttk.Button(self.input_frame, text="START", command=self.start_data_collection)
        self.start_button.grid(row=11, column=0, columnspan=2, pady=10)
        self.stop_button = ttk.Button(self.input_frame, text="STOP", command=self.on_stop)
        self.stop_button.grid(row=12, column=0, columnspan=2, pady=10)
        ttk.Label(self.input_frame, text="NOTE: This will only stop data collection during the frequency sweep.").grid(row=13, column=0, columnspan=2, pady=2)
        self.cancel_button = ttk.Button(self.input_frame, text="RESET", command=self.on_cancel)
        self.cancel_button.grid(row=15, column=0, columnspan=2, pady=10)
        ttk.Label(self.input_frame, text="NOTE: This will reset the program and no data will be saved.").grid(row=16, column=0, columnspan=2, pady=2)

        # Message feed to display live updates during measurements
        self.message_feed = tk.Text(self.input_frame, wrap="word", height=10, width=50)
        self.message_feed.grid(row=14, column=0, columnspan=3, padx=5, pady=5)

        # New button for zeroing the power sensor manually
        self.zero_button = ttk.Button(self.input_frame, text="Zero Power Sensor", command=self.start_zeroing)
        self.zero_button.grid(row=17, column=0, columnspan=2, pady=10)
        ttk.Label(self.input_frame, text="Click to zero the power meter").grid(row=18, column=0, columnspan=2)
        ttk.Label(self.input_frame, text="NOTE: VOA output must be disabled! The PD should not be receiving any light.").grid(row=19, column=0, columnspan=2)

    def setup_plots(self):
        # Detect screen resolution and set parameters accordingly.
        screen_width = self.root.winfo_screenwidth()
        if screen_width <= 1280:
            fig_size = (9, 7)
            title_font = {'size': 10, 'weight': 'bold'}
            label_font = {'size': 8, 'weight': 'bold'}
            tick_font_size = 8
            hspace = 0.49
            wspace = 0.49
        else:
            fig_size = (12, 9)
            title_font = {'size': 14, 'weight': 'bold'}
            label_font = {'size': 12, 'weight': 'bold'}
            tick_font_size = 12
            hspace = 0.4
            wspace = 0.4

        # Create the figure without using constrained_layout.
        self.fig = Figure(figsize=fig_size)
        # Reserve the top area (15% of the figure) for annotations.
        self.fig.subplots_adjust(top=0.85, bottom=0.1, left=0.13, right=0.99, hspace=hspace, wspace=wspace)
        
        # Create a 2x2 grid of subplots.
        self.ax1 = self.fig.add_subplot(221)
        self.ax3 = self.fig.add_subplot(222)
        self.ax4 = self.fig.add_subplot(223)
        self.ax5 = self.fig.add_subplot(224)
        # Create twin axis for the first subplot.
        self.ax2 = self.ax1.twinx()

        # Set titles and labels using dynamic fonts.
        self.ax1.set_title('Beat Frequency vs Step Number', fontdict=title_font)
        self.ax1.set_xlabel('Step Number', fontdict=label_font)
        self.ax1.set_ylabel('Beat Frequency (GHz)', fontdict=label_font)
        self.ax2.set_ylabel('Laser 4 Wavelength (nm)', fontdict=label_font)
        self.ax3.set_title('Raw RF Power vs Beat Frequency', fontdict=title_font)
        self.ax3.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
        self.ax3.set_ylabel('Raw RF Power (dBm)', fontdict=label_font)
        self.ax4.set_title('Measured Photocurrent vs Beat Frequency', fontdict=title_font)
        self.ax4.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
        self.ax4.set_ylabel('Photocurrent (mA)', fontdict=label_font)
        self.ax5.set_title('Calibrated RF Power vs Beat Frequency', fontdict=title_font)
        self.ax5.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
        self.ax5.set_ylabel('Calibrated RF Power (dBm)', fontdict=label_font)

        # Set tick label sizes.
        self.ax1.tick_params(axis='y', labelsize=tick_font_size, labelcolor='tab:blue')
        self.ax2.tick_params(axis='y', labelsize=tick_font_size, labelcolor='tab:red')
        self.ax3.tick_params(axis='y', labelsize=tick_font_size, labelcolor='tab:blue')
        self.ax4.tick_params(axis='y', labelsize=tick_font_size, labelcolor='tab:blue')
        self.ax5.tick_params(axis='y', labelsize=tick_font_size, labelcolor='tab:blue')

        # Create plot lines and markers.
        self.line1, = self.ax1.plot([], [], linestyle='-', color='tab:blue')
        self.markers1, = self.ax1.plot([], [], 'o', color='tab:blue')
        self.line2, = self.ax2.plot([], [], linestyle='--', color='tab:red')
        self.markers2, = self.ax2.plot([], [], 'x', color='tab:red')
        self.line3, = self.ax3.plot([], [], linestyle='-', color='tab:blue')
        self.markers3, = self.ax3.plot([], [], 'o', color='tab:blue')
        self.line4, = self.ax4.plot([], [], linestyle='-', color='tab:blue')
        self.markers4, = self.ax4.plot([], [], 'o', color='tab:blue')
        self.line5, = self.ax5.plot([], [], linestyle='-', color='tab:blue')
        self.markers5, = self.ax5.plot([], [], 'o', color='tab:blue')

        # Embed the figure into the Tkinter plot frame.
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.plot_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.plot_frame)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Add hover functionality with mplcursors.
        mplcursors.cursor(self.markers1, hover=mplcursors.HoverMode.Transient)
        mplcursors.cursor(self.markers3, hover=mplcursors.HoverMode.Transient)
        mplcursors.cursor(self.markers4, hover=mplcursors.HoverMode.Transient)
        mplcursors.cursor(self.markers5, hover=mplcursors.HoverMode.Transient)

        # Store dynamic fonts for later use in save_data.
        self.dynamic_title_font = title_font
        self.dynamic_label_font = label_font




    def select_s2p_file(self):
        """
        Open a file dialog for selecting a .s2p file (for RF Link Loss calibration).
        """
        file_path = filedialog.askopenfilename(
            title="Select .s2p File",
            filetypes=[("s2p files", "*.s2p"), ("All files", "*.*")]
        )
        if file_path:
            self.s2p_file_var.set(file_path)

    def select_excel_file(self):
        """
        Open a file dialog for selecting an Excel (.xlsx) file (for RF Probe Loss calibration).
        """
        file_path = filedialog.askopenfilename(
            title="Select .xlsx File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file_var.set(file_path)

    def update_message_feed(self, message: str):
        """
        Update the message feed in the GUI with the given message.
        This is used to provide real-time feedback during the measurement process.
        """
        self.message_feed.insert(tk.END, message + "\n")
        self.message_feed.see(tk.END)
        self.root.update_idletasks()

    def open_instruments(self):
        """
        Create VISA adapters with the connected equipment.
        If connection fails, report the error to the message feed.
        """
        try:
            self.ecl_adapter = self.rm.open_resource(self.ecl_adapter_GPIB)
            self.wavelength_meter = self.rm.open_resource(self.wavelength_meter_GPIB)
            self.spectrum_analyzer = self.rm.open_resource(self.spectrum_analyzer_GPIB)
            self.keithley = self.rm.open_resource(self.keithley_GPIB)
            self.power_sensor = self.rm.open_resource(self.power_sensor_GPIB)
            self.voa = self.rm.open_resource(self.voa_GPIB)

            self.power_sensor.timeout = 15000
            self.ecl_adapter.timeout = 10000
            self.wavelength_meter.timeout = 10000
            self.spectrum_analyzer.timeout = 10000
            self.keithley.timeout = 10000
            self.voa.timeout = 10000
        
            self.update_message_feed("Successfully connected to all VISA devices.")
        except Exception as e:
            self.update_message_feed(f"Error connecting to VISA devices: {e}")

    def close_instruments(self):
        try:
            if self.ecl_adapter is not None:
                self.ecl_adapter.close()
                self.ecl_adapter = None
            if self.wavelength_meter is not None:
                self.wavelength_meter.close()
                self.wavelength_meter = None
            if self.spectrum_analyzer is not None:
                self.spectrum_analyzer.close()
                self.spectrum_analyzer = None
            if self.keithley is not None:
                self.keithley.close()
                self.keithley = None
            if self.power_sensor is not None:
                self.power_sensor.close()
                self.power_sensor = None
            if self.voa is not None:
                self.voa.close()
                self.voa = None
            self.update_message_feed("Instruments closed successfully.")
        except Exception as e:
            self.update_message_feed(f"Error closing instruments: {e}")

    def measure_peak_frequency(self):
        """
        Measure the peak frequency using the spectrum analyzer.
        (Repeated measurements are taken and the minimum value is returned.)
        """
        try:
            self.spectrum_analyzer.write('MKPK HI')
            time.sleep(0.1)
            peak_freq_1 = self.spectrum_analyzer.query('MKF?')
            time.sleep(0.1)
            self.spectrum_analyzer.write('MKPK HI')
            time.sleep(0.1)
            peak_freq_2 = self.spectrum_analyzer.query('MKF?')
            time.sleep(0.1)
            self.spectrum_analyzer.write('MKPK HI')
            time.sleep(0.1)
            peak_freq_3 = self.spectrum_analyzer.query('MKF?')
            peak_freq = min(peak_freq_1, peak_freq_2, peak_freq_3)

            self.spectrum_analyzer.write(":SENS:FREQ:SPAN 50GHz") # reset span to full span for next measurement
            self.spectrum_analyzer.write(":SYSTem:LOCal")
            return float(peak_freq) / 1e9  # Convert Hz to GHz
        
        except Exception as e:
            self.update_message_feed(f"Error measuring peak frequency: {e}")
            return None

    def measure_wavelength_beat(self):
        """
        Measure the beat frequency using the wavelength meter.
        Initiates a single measurement and then calculates the difference between two frequency readings.
        If the returned beat frequency is below a valid threshold (e.g. 50 GHz), or if nothing is returned,
        return None so that the ESA measurement is used instead.
        """
        try:
            self.wavelength_meter.write(":INIT:IMM")
            time.sleep(1.5)
            result = self.wavelength_meter.query(":CALC3:DATA? FREQuency").strip()
            # If the query returns an empty string, return None immediately.
            if not result:
                return None
            data = result.split(',')
            freqs = [float(f) for f in data]
            beat_node = min(freqs)  # assuming the reference delta is 0
            self.wavelength_meter.write(":SYSTem:LOCal")
            beat_val = abs(beat_node / 1e9)  # Convert to GHz
            if beat_val < 50:  # Threshold check
                return None
            return beat_val
        except Exception as e:
            self.update_message_feed(f"Error measuring beat frequency with wavelength meter: {e}")
            return None

    def zero_power_sensor(self):
        try:
            output_state = self.voa.query(":OUTPut:STATe?").strip()
            if int(output_state) == 1:
                self.update_message_feed("VOA is currently enabled, disable before zeroing power meter...")
                raise Exception("VOA is enabled")
            self.update_message_feed("Zeroing power sensor...")
            # Anritsu ML2437A Power Meter:
            self.power_sensor.write('ZERO A')
            time.sleep(10)  # wait 10 seconds for the zeroing process to complete
            self.update_message_feed("Power sensor zeroing completed.")
            self.voa.write(":SYSTem:LOCal")
        except Exception as e:
            self.update_message_feed(f"Error during power sensor zeroing: {e}")

    def measure_rf_power(self, beat_freq, max_attempts=3):
        """
        Safely measure RF power at beat_freq. Returns dBm or None if all attempts time out.
        """
        for attempt in range(1, max_attempts+1):
            try:
                self.power_sensor.write(f"CFFRQ A,{beat_freq}E9")
                self.power_sensor.write('CFSRC A,FREQ')
                
                time.sleep(0.2)

                readings = []
                for _ in range(5):
                    self.power_sensor.write('STA 1')
                    time.sleep(0.1)  # Allow time for the measurement to stabilize

                    reading = self.power_sensor.query('O 1')
                    pwr_dbm = float(reading.strip())

                    # Append dBm reading for averaging
                    readings.append(pwr_dbm)
                    time.sleep(0.1)

                avg_pow = sum(readings) / len(readings)
                return avg_pow

            except pyvisa.errors.VisaIOError:
                self.update_message_feed(
                    f"Power sensor timeout (attempt {attempt}/{max_attempts}), retrying..."
                )
                time.sleep(0.5)

        self.update_message_feed(
            "Power sensor unavailable after retries—continuing sweep without RF data."
        )
        return None


    def start_zeroing(self):
        # Launch the zeroing process in a separate thread
        threading.Thread(target=self.zero_power_sensor, daemon=True).start()

    def set_laser_wavelength(self, channel: int, wavelength: float):
        """
        Set the laser wavelength for the specified channel.
        (This sends the command to the ECL laser via VISA.)
        """
        self.update_message_feed(f"Setting laser {channel} wavelength to {wavelength:.3f} nm...")
        self.ecl_adapter.write(f"CH{channel}:L={wavelength:.3f}")

    def read_excel_data(self, filepath: str):
        """
        Read the Excel file containing RF probe loss data.
        Returns two numpy arrays: one for frequency and one for loss.
        """
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        frequency = []
        loss = []
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            frequency.append(row[0])
            loss.append(row[1])
        return np.array(frequency), np.array(loss)

    def read_s2p_file(self, filepath: str):
        """
        Read the .s2p file containing network analyzer data.
        Extracts frequency and S-parameter (S12 and S21) data and returns the averaged loss.
        """
        frequencies = []
        s12 = []
        s21 = []
        with open(filepath, 'r') as file:
            lines = file.readlines()
        freq_unit = 'hz'
        data_format = 'db'
        for line in lines:
            if line.startswith('#'):
                parts = line.split()
                for part in parts:
                    if part.lower() in ['hz', 'khz', 'mhz', 'ghz']:
                        freq_unit = part.lower()
                    elif part.lower() in ['ri', 'db', 'ma']:
                        data_format = part.upper()
                continue
            if not line.startswith('!') and not line.startswith('#'):
                values = re.split(r'\s+', line.strip())
                if len(values) >= 9:
                    frequencies.append(float(values[0]))
                    s21.append(float(values[3]))  # S21 in dB
                    s12.append(float(values[5]))  # S12 in dB
        frequencies = np.array(frequencies)
        s_avg = (np.array(s12) + np.array(s21)) / 2
        if freq_unit == 'khz':
            frequencies = frequencies / 1e6
        elif freq_unit == 'mhz':
            frequencies = frequencies / 1e3
        elif freq_unit == 'hz':
            frequencies = frequencies / 1e9
        return frequencies, s_avg

    def custom_linear_interpolation(self, x, y, x_new):
        x = np.asarray(x)
        y = np.asarray(y)
        x_new = np.asarray(x_new)
        y_new = np.zeros_like(x_new)
        for i, xi in enumerate(x_new):
            if xi <= x[0]:
                y_new[i] = y[0]
            elif xi >= x[-1]:
                y_new[i] = y[-1]
            else:
                k = np.searchsorted(x, xi) - 1
                x_k = x[k]
                x_k1 = x[k + 1]
                y_k = y[k]
                y_k1 = y[k + 1]
                # Avoid division by zero if consecutive x values are identical
                if x_k1 - x_k == 0:
                    y_new[i] = y_k
                else:
                    y_new[i] = y_k + (y_k1 - y_k) * (xi - x_k) / (x_k1 - x_k)
        return y_new


    def calculate_calibrated_rf(self, powers, beat_freqs_pow, s2p_filename=None, excel_filename=None):
        """
        Calculate the calibrated RF power by applying RF losses from:
         - A network analyzer (.s2p file) for RF link loss.
         - An Excel file for RF probe loss.
        Returns the calibrated RF power and the individual loss components.
        """
        calibrated_rf = np.array(powers)  # Initialize with raw RF power
        rf_loss = np.zeros_like(calibrated_rf)
        rf_probe_loss = np.zeros_like(calibrated_rf)
        rf_link_loss = np.zeros_like(calibrated_rf)
        # Apply calibration from .s2p file
        if s2p_filename:
            try:
                frequencies, s_avg = self.read_s2p_file(s2p_filename)
                interpolated_loss = self.custom_linear_interpolation(frequencies, s_avg, beat_freqs_pow)
                interpolated_loss = np.array(interpolated_loss)
                rf_loss += np.abs(interpolated_loss.real)
                rf_probe_loss += np.abs(interpolated_loss.real)
                calibrated_rf += np.abs(interpolated_loss.real)
                calibrated_rf = np.round(calibrated_rf, 2)
            except Exception as e:
                self.update_message_feed(f"Error processing S2P file: {e}")
        # Apply calibration from Excel file
        if excel_filename:
            try:
                probe_loss_frequency, probe_loss = self.read_excel_data(excel_filename)
                interpolated_probe_loss = self.custom_linear_interpolation(probe_loss_frequency, probe_loss, beat_freqs_pow)
                interpolated_probe_loss = np.array(interpolated_probe_loss)
                rf_loss += np.abs(interpolated_probe_loss)
                rf_link_loss += np.abs(interpolated_probe_loss)
                calibrated_rf += np.abs(interpolated_probe_loss)
                calibrated_rf = np.round(calibrated_rf, 2)
            except Exception as e:
                self.update_message_feed(f"Error processing Excel file: {e}")
        return calibrated_rf, rf_loss, rf_probe_loss, rf_link_loss

    def validate_inputs(self) -> bool:
        """
        Check that all GUI input fields (numeric ones) are valid.
        This adds basic type checking so the program does not crash due to bad input.
        """
        try:
            float(self.laser_3_var.get())
            float(self.laser_4_var.get())
            float(self.start_freq_var.get())
            float(self.end_freq_var.get())
            int(self.num_steps_var.get())
            float(self.delay_var.get())
            return True
        except Exception as e:
            messagebox.showerror("Input Error", f"Invalid input: {e}")
            return False

    def start_data_collection(self):
        """
        Begin data collection by validating inputs and launching the measurement process
        in a separate thread so that the GUI remains responsive.
        """
        if not self.validate_inputs():
            return
        # Store the thread reference so we can join it later
        self.measurement_thread = threading.Thread(target=self.data_collection, daemon=True)
        self.measurement_thread.start()

    def toggle_resume_button(self):
        if self.pause_after_search_var.get():
            self.resume_button.config(state=tk.NORMAL)
        else:
            self.resume_button.config(state=tk.DISABLED)

    def resume_after_pause(self):
        self.resume_button.config(state=tk.DISABLED)
        self.update_message_feed("Resuming measurement after pause.")
        self.pause_event.set()

    def data_collection(self):
        """
        Perform the complete data collection process:
         - Open instruments and wait for stabilization.
         - Set initial laser wavelengths.
         - Perform an automatic beat frequency search loop (if enabled) to bring the system near the target frequency.
         - Conduct the frequency sweep and measure beat frequency, photocurrent, and RF power.
         - Handle occasional errors and measurement retries.
         - Finally, perform calibration calculations.
         
         (Many inline comments explain each step, as in the original code.)
        """

        input_window = tk.Toplevel(self.root)
        input_window.title("Save Data Inputs")
        input_window.geometry("300x200")
        
        ttk.Label(input_window, text="Device Number:").grid(row=0, column=0, padx=10, pady=10)
        device_num_var = tk.StringVar()
        device_num_entry = ttk.Entry(input_window, textvariable=device_num_var)
        device_num_entry.grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(input_window, text="Comments:").grid(row=1, column=0, padx=10, pady=10)
        comment_var = tk.StringVar()
        comment_entry = ttk.Entry(input_window, textvariable=comment_var)
        comment_entry.grid(row=1, column=1, padx=10, pady=10)
        
        def choose_file_and_close():
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            if not file_path:
                return  # Optionally, handle cancelation here
            # Store the user inputs in instance variables for later use
            self.device_num = device_num_var.get().strip()
            self.user_comment = comment_var.get().strip().upper()
            self.save_file_path = file_path
            self.excel_file_path = file_path.replace(".txt", ".xlsx")
            self.plot_file_path = file_path.rsplit('.', 1)[0] + '.png'
            input_window.destroy()
        
        confirm_button = ttk.Button(input_window, text="Confirm and Choose Save Location", command=choose_file_and_close)
        confirm_button.grid(row=2, column=0, columnspan=2, pady=20)
        
        # Wait for the user to provide the inputs and choose a save location
        self.root.wait_window(input_window)
        # ===============================================================
        self.stop_event.clear()
        self.data_ready_event.clear()

        start_time = time.time()
        try:
            # --- Get user inputs and initialize measurement parameters ---
            laser_3_WL = self.laser_3_var.get()
            laser_4_WL = self.laser_4_var.get()
            num_steps = self.num_steps_var.get()
            delay = self.delay_var.get()
            start_freq = self.start_freq_var.get()
            end_freq = self.end_freq_var.get()
            enable_search = self.enable_search_var.get()
            freq_threshold = 0.5  # Note: values below 0.5 GHz are less likely to work
            excel_filename = self.excel_file_var.get()
            s2p_filename = self.s2p_file_var.get()

            # Set the laser wavelengths and power
            self.set_laser_wavelength(3, laser_3_WL)
            self.set_laser_wavelength(4, laser_4_WL)

            # Wait for the lasers to stabilize
            self.update_message_feed("Waiting for the lasers to stabilize...")
            time.sleep(10)

            # Initialize frequencies: set reference frequency to laser 3
            c = 299792458  # Speed of light in m/s
            laser_3_freq = c / (laser_3_WL * 1e-9)
            self.wavelength_meter.write(":CALCulate3:PRESet")
            time.sleep(1)
            self.wavelength_meter.write("*OPC?")
            if self.wavelength_meter.read().strip() == "1":
                print("CALCulate3 states cleared.")
            else:
                print("Warning: Operation did not complete as expected.")

            self.wavelength_meter.write(":CALCulate3:DELTa:REFerence:WAVelength MIN")
            time.sleep(1)

            self.wavelength_meter.write(":CALCulate3:DELTa:WAVelength ON")
            time.sleep(1)
            self.wavelength_meter.write("*OPC?")
            if self.wavelength_meter.read().strip() == "1":
                print("Delta wavelength mode successfully turned ON.")
            else:
                print("Delta wavelength mode command did not complete as expected.")
            time.sleep(1)

            # Additional variables to track consecutive increases
            consecutive_increases = 0
            max_consecutive_increases = 3

            # Measure initial beat frequency using both instruments
            wl_meter_beat_freq = self.measure_wavelength_beat()
            esa_beat_freq = self.measure_peak_frequency()
            if wl_meter_beat_freq is None:
                wl_meter_beat_freq = esa_beat_freq
            current_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq
            last_beat_freq = None

            # --- AUTO START FREQUENCY SEARCH LOOP ---
            if enable_search:
                self.update_message_feed("RUNNING AUTOMATIC START FREQUENCY SEARCH LOOP...")
                while current_freq >= 1:
                    if self.stop_event.is_set():
                        self.update_message_feed("Data collection stopped by user.")
                        return
                    laser_4_freq = c / (laser_4_WL * 1e-9)
                    wl_meter_beat_freq = self.measure_wavelength_beat()
                    esa_beat_freq = self.measure_peak_frequency()
                    if wl_meter_beat_freq is None:
                        wl_meter_beat_freq = esa_beat_freq
                    if wl_meter_beat_freq is None or esa_beat_freq is None:
                        self.update_message_feed("Issue reading from WLM and ESA, updating small jump in laser...")
                        laser_4_freq = c / (laser_4_WL * 1e-9)
                        laser_4_new_freq = laser_4_freq - (0.2 * 1e9)
                        laser_4_WL = (c / laser_4_new_freq) * 1e9
                        self.set_laser_wavelength(4, laser_4_WL)
                        time.sleep(3)
                        continue
                    current_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq
                    if last_beat_freq is not None and current_freq > last_beat_freq:
                        if self.stop_event.is_set():
                            self.update_message_feed("Data collection stopped by user.")
                            return
                        if current_freq >= 1:
                            consecutive_increases += 1
                            if consecutive_increases >= max_consecutive_increases:
                                self.update_message_feed("Start frequency search overshot the threshold, restarting...")
                                # Reset calibration variables
                                laser_4_WL = laser_3_WL - 2  # Reset to an initial offset
                                self.set_laser_wavelength(4, laser_4_WL)
                                consecutive_increases = 0
                                last_beat_freq = None
                                time.sleep(15)
                                wl_meter_beat_freq = self.measure_wavelength_beat()
                                esa_beat_freq = self.measure_peak_frequency()
                                if wl_meter_beat_freq is None:
                                    wl_meter_beat_freq = esa_beat_freq
                                if wl_meter_beat_freq is None or esa_beat_freq is None:
                                    continue
                                current_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq
                                continue
                        elif current_freq < 1:
                            break  # Near 0 beat frequency
                    else:
                        consecutive_increases = 0

                    if self.stop_event.is_set():
                        self.update_message_feed("Data collection stopped by user.")
                        return

                    if wl_meter_beat_freq >= 50 and wl_meter_beat_freq < 1000:
                        self.update_message_feed(f"Beat Frequency (Wavelength Meter): {wl_meter_beat_freq} GHz")
                        laser_4_new_freq = laser_4_freq - (wl_meter_beat_freq * 0.67 * 1e9)
                        laser_4_WL = (c / laser_4_new_freq) * 1e9
                        if 1540 < laser_4_WL < 1660:
                            self.set_laser_wavelength(4, laser_4_WL)
                        else:
                            self.update_message_feed(f"New wavelength out of bounds: {laser_4_WL:.3f} nm")
                            return
                    elif esa_beat_freq < 50 and (wl_meter_beat_freq < 50 or wl_meter_beat_freq > 10000):
                        self.update_message_feed(f"Beat Frequency (ESA): {round(esa_beat_freq,2)} GHz")
                        if esa_beat_freq > 3:
                            if last_beat_freq is not None and last_beat_freq < 1:
                                laser_4_new_freq = laser_4_freq - (0.2 * 1e9)
                            else:
                                laser_4_new_freq = laser_4_freq - (esa_beat_freq * 0.67 * 1e9)
                        elif 1.5 < esa_beat_freq <= 3:
                            laser_4_new_freq = laser_4_freq - (0.5 * 1e9)
                        elif 1 <= esa_beat_freq <= 1.5:
                            laser_4_new_freq = laser_4_freq - (0.2 * 1e9)
                        elif esa_beat_freq < 1:
                            laser_4_new_freq = laser_4_freq - (0.1 * 1e9)
                        laser_4_WL = (c / laser_4_new_freq) * 1e9
                        if 1540 < laser_4_WL < 1660:
                            self.set_laser_wavelength(4, laser_4_WL)
                        else:
                            self.update_message_feed(f"New wavelength out of bounds: {laser_4_WL:.3f} nm")
                            return
                    last_beat_freq = current_freq
                    time.sleep(3)
                # After loop, attempt a small jump to overcome ESA measurement issues near 0 GHz
                self.update_message_feed("Attempting small jump over ESA issues near 0 GHz...")
                laser_4_freq = c / (laser_4_WL * 1e-9)
                laser_4_new_freq = laser_4_freq - (1 * 1e9)
                laser_4_WL = (c / laser_4_new_freq) * 1e9
                self.set_laser_wavelength(4, laser_4_WL)
                time.sleep(3)
                wl_meter_beat_freq = self.measure_wavelength_beat()
                esa_beat_freq = self.measure_peak_frequency()
                if wl_meter_beat_freq is None:
                    wl_meter_beat_freq = esa_beat_freq
                current_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq
                self.update_message_feed(f"Current Beat Frequency: {round(current_freq,2)} GHz")

                # Instead of a single if, use a while loop for the second jump if current_freq > 2
                max_attempts = 5
                attempt = 0
                while current_freq is not None and current_freq > 10 and attempt < max_attempts:
                    self.update_message_feed("Attempting second small jump over ESA issues near 0 GHz...")
                    laser_4_freq = c / (laser_4_WL * 1e-9)
                    laser_4_new_freq = laser_4_freq - (0.4 * 1e9)
                    laser_4_WL = (c / laser_4_new_freq) * 1e9
                    self.set_laser_wavelength(4, laser_4_WL)
                    time.sleep(3)
                    wl_meter_beat_freq = self.measure_wavelength_beat()
                    esa_beat_freq = self.measure_peak_frequency()
                    if wl_meter_beat_freq is None:
                        wl_meter_beat_freq = esa_beat_freq
                    current_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq
                    self.update_message_feed(f"Current Beat Frequency: {round(current_freq,2)} GHz")
                    attempt += 1

                 # --- SECOND LOOP: Adjust laser 4 to reach the desired starting frequency ---
                if start_freq > 1:
                    self.update_message_feed("Adjusting laser 4 to reach starting beat frequency after passing 0...")
                    update_laser = True
                    while abs(current_freq - start_freq) > freq_threshold:
                        if self.stop_event.is_set():
                            self.update_message_feed("Data collection stopped by user.")
                            break

                        # Re-measure beat frequency using both instruments
                        wl_meter_beat_freq = self.measure_wavelength_beat()
                        esa_beat_freq = self.measure_peak_frequency()
                        if wl_meter_beat_freq is None:
                            wl_meter_beat_freq = esa_beat_freq
                        if wl_meter_beat_freq is None or esa_beat_freq is None:
                            continue

                        current_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 45 and wl_meter_beat_freq < 1000) else esa_beat_freq
                        self.update_message_feed(f"Current Beat Frequency: {round(current_freq,2)} GHz")

                        # Calculate the current frequency of laser 4 and update it by half the difference
                        laser_4_freq = c / (laser_4_WL * 1e-9)
                        laser_4_new_freq = laser_4_freq - ((abs(start_freq - current_freq) * 1e9)) / 2
                        laser_4_WL = (c / laser_4_new_freq) * 1e9

                        if abs(current_freq - start_freq) <= freq_threshold:
                            update_laser = False

                        if update_laser:
                            if 1540 < laser_4_WL < 1660:
                                self.set_laser_wavelength(4, laser_4_WL)
                            else:
                                self.update_message_feed(f"New wavelength for laser 4 is out of bounds: {laser_4_WL:.3f} nm")
                                return

                        time.sleep(3)
                        last_beat_freq = current_freq

                # Final measurement after loop finishes
                time.sleep(5)
                wl_meter_beat_freq = self.measure_wavelength_beat()
                esa_beat_freq = self.measure_peak_frequency()
                if wl_meter_beat_freq is None:
                    wl_meter_beat_freq = esa_beat_freq
                current_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq
                self.update_message_feed(f"Final Beat Frequency: {round(current_freq,2)} GHz")
                enable_search = False  # Disable the search after reaching the starting frequency

            if self.pause_after_search_var.get():
                self.update_message_feed("PAUSED after initial search. Click ‘Resume Measurements’ to continue.")
                # clear any previous resume signal
                self.pause_event.clear()
                # block this thread until resume_event is set by your button
                self.pause_event.wait()

            # --- BEGIN DATA COLLECTION LOOP ---
            # Calculate step size for laser 4 frequency adjustment
            laser_4_step = (end_freq - start_freq) / num_steps
            self.update_message_feed("BEGINNING MEASUREMENT LOOP...")
            start_time_sweep = time.time()
            last_beat_freq = current_freq

            # Get initial photocurrent from Keithley (convert to mA)
            response = self.keithley.query(":MEASure:CURRent?")
            initial_current_values = response.split(',')
            if len(initial_current_values) > 1:
                initial_current = float(initial_current_values[1]) * 1000
                initial_current = round(initial_current, 3)
            self.keithley.write(":SYSTem:LOCal")

            self.looping = True
            for step in range(num_steps):
                if self.stop_event.is_set():
                    self.update_message_feed("Data collection stopped by user.")
                    self.looping = False
                    time_end = time.time()
                    sweep_run_time = time_end - start_time_sweep
                    total_run_time = time_end - start_time
                    beat_freqs, powers, photo_currents, p_actuals = zip(*self.beat_freq_and_power)
                    self.calibrated_rf, self.rf_loss, self.rf_probe_loss, self.rf_link_loss = self.calculate_calibrated_rf(
                        powers, beat_freqs, s2p_filename=s2p_filename, excel_filename=excel_filename
                    )
                    self.data_ready_event.set()
                    break

                # Choose measurement method based on previous beat frequency
                wl_meter_beat_freq = self.measure_wavelength_beat()
                esa_beat_freq = self.measure_peak_frequency()
                if wl_meter_beat_freq is None:
                    wl_meter_beat_freq = esa_beat_freq
                beat_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq

                self.update_message_feed(f"Step {step + 1} of {num_steps}")

                # For early steps near low start frequencies, adjust laser 4 more cautiously
                if step < 2 and start_freq < 5 and beat_freq > 15:
                    laser_4_freq = c / (laser_4_WL * 1e-9)
                    laser_4_new_freq = laser_4_freq - (0.3 * 1e9)
                    laser_4_WL = (c / laser_4_new_freq) * 1e9
                    self.set_laser_wavelength(4, laser_4_WL)
                    time.sleep(delay)
                    wl_meter_beat_freq = self.measure_wavelength_beat()
                    esa_beat_freq = self.measure_peak_frequency()
                    if wl_meter_beat_freq is None:
                        wl_meter_beat_freq = esa_beat_freq
                    beat_freq = wl_meter_beat_freq if (wl_meter_beat_freq > 50 and wl_meter_beat_freq < 1000) else esa_beat_freq
                    if beat_freq is None:
                        continue

                # --- MEASURE CURRENT FROM KEITHLEY ---
                response = self.keithley.query(":MEASure:CURRent?")
                current_values = response.split(',')
                if len(current_values) > 1:
                    current = float(current_values[1]) * 1000  # Convert to mA
                    current = round(current, 3)
                self.keithley.write(":SYSTem:LOCal")

                # --- MEASURE VOA and RF power sensor data with retry loop ---
                max_attempts = 3
                attempts = 0
                success = False
                p_actual = self.voa.query('READ:POW?')
                p_actual = round(float(p_actual), 3)
                self.voa.write('SYST:LOC')
                
                output_dbm = self.measure_rf_power(beat_freq)
                if output_dbm is None:
                    # record placeholder (or skip)
                    self.beat_freq_and_power.append(
                        (beat_freq, float('nan'), current, p_actual)
                    )
                else:
                    power = round(output_dbm, 2)
                    self.beat_freq_and_power.append(
                        (beat_freq, power, current, p_actual)
                    )

                self.update_message_feed(f"Beat Frequency: {round(beat_freq,2)} GHz")
                self.update_message_feed(f"Measured Photocurrent: {current} mA")
                self.update_message_feed(f"Raw RF Power: {output_dbm} dBm")
                self.steps.append(step + 1)
                self.beat_freqs.append(beat_freq)
                self.laser_4_wavelengths.append(laser_4_WL)
                self.rf_loss.append(0)  # Placeholder for RF loss
                self.calibrated_rf.append(output_dbm)  # Placeholder for calibrated RF power
                self.photo_currents.append(current)
                self.powers.append(output_dbm)

                # Update laser 4 wavelength for the next step
                laser_4_freq = c / (laser_4_WL * 1e-9)
                laser_4_new_freq = laser_4_freq - (laser_4_step * 1e9)
                laser_4_WL = (c / laser_4_new_freq) * 1e9
                self.set_laser_wavelength(4, laser_4_WL)
                last_beat_freq = beat_freq
                self.data_ready_event.set()
                time.sleep(delay)

            self.update_message_feed("Data collection completed.")
            self.looping = False
            time_end = time.time()
            sweep_run_time = time_end - start_time_sweep
            total_run_time = time_end - start_time
            beat_freqs, powers, photo_currents, p_actuals = zip(*self.beat_freq_and_power)
            self.p_actuals = list(p_actuals)
            self.calibrated_rf, self.rf_loss, self.rf_probe_loss, self.rf_link_loss = self.calculate_calibrated_rf(
                powers, beat_freqs, s2p_filename=s2p_filename, excel_filename=excel_filename
            )
            self.data_ready_event.set()

            # After data collection, prompt user for additional inputs and save data
            volt = self.keithley.query(':SOUR:VOLT:LEV:IMM:AMPL?').strip()
            self.keithley_voltage = volt

            self.root.after(0, lambda: self.save_data(sweep_run_time, total_run_time))
            return
        except Exception as e:
            self.update_message_feed(f"Error in data collection: {e}")
            self.reset_program()

    def save_data(self, sweep_run_time, total_run_time):
        """
        Runs in the GUI thread:  
        - Annotate the figure  
        - Draw the canvas *lightly*  
        - Spawn a background thread for all heavy file I/O  
        """
        # --- 1) annotation (unchanged) ---
        device_num    = self.device_num
        user_comment  = self.user_comment
        file_path     = self.save_file_path
        plot_file_path= self.plot_file_path

        comments = [
            f"Device Number: {device_num}",
            f"Comments: {user_comment}",
            f"Date: {time.strftime('%m/%d/%Y')}",
            f"Time: {time.strftime('%H:%M:%S')}",
            f"Frequency Sweep Run Time: {sweep_run_time:.2f} s",
            f"Total Run Time: {total_run_time:.2f} s",
            f"Keithley Voltage: {self.keithley_voltage} V",
            f"Excel Loss File: {self.excel_file_var.get() or 'None'}",
            f"S2P Loss File: {self.s2p_file_var.get() or 'None'}"
        ]


            # 1) annotate & draw on main thread
        self._annotate_and_draw(comments)

        # 2) save the figure PNG *right here* (main thread)
        try:
            self.fig.savefig(plot_file_path)
        except Exception as e:
            self.update_message_feed(f"Plot save failed: {e}")

        # 3) hand off only the text + Excel work to a background thread
        threading.Thread(
            target=self._save_data_io,
            args=(file_path, plot_file_path, sweep_run_time, total_run_time),
            daemon=True
        ).start()


    def _annotate_and_draw(self, comments):
        """
        Runs on the GUI thread: update ticks, add text annotations and redraw.
        """
        # ---- 1) update the y-ticks on ax3 and ax5 ----
        # raw RF power axis (ax3)
        min_power = min(self.powers)
        max_power = max(self.powers)
        yticks = np.arange(
            np.floor(min_power/3)*3,
            np.ceil(max_power/3)*3 + 3,
            3
        )
        self.ax3.set_yticks(yticks)
        self.ax3.set_ylim(min(yticks), max(yticks))

        # calibrated RF power axis (ax5)
        min_calibrated_rf = min(self.calibrated_rf)
        max_calibrated_rf = max(self.calibrated_rf)
        yticks2 = np.arange(
            np.floor(min_calibrated_rf/3)*3,
            np.ceil(max_calibrated_rf/3)*3 + 3,
            3
        )
        self.ax5.set_yticks(yticks2)
        self.ax5.set_ylim(min(yticks2), max(yticks2))

        # ---- 2) place your comments text ----
        n = len(comments)
        n_left = math.ceil(n/2)
        for i, text in enumerate(comments):
            x = 0.3 if i < n_left else 0.7
            y = 0.94 - ((i if i < n_left else i-n_left)*0.02)
            self.fig.text(x, y, text, wrap=True, ha='center', fontsize=10)

        # ---- 3) one non-blocking redraw ----
        self.canvas.draw_idle()


    def _save_data_io(self, file_path, plot_file_path, sweep_run_time, total_run_time):
        """
        Runs in a background thread:  
        - Save figure PNG  
        - Write text file  
        - Build & save Excel workbook  
        """

        # 2) write the text file
        try:
            with open(file_path, 'w') as f:
                f.write("DEVICE NUMBER: " + str(self.device_num) + "\n")
                f.write("COMMENTS: " + self.user_comment + "\n")
                f.write("KEITHLEY VOLTAGE: " + str(self.keithley_voltage) + " V" + "\n")
                f.write("FREQUENCY SWEEP RUN TIME: " + f"{sweep_run_time:.2f}" + " s" + "\n")
                f.write("TOTAL RUN TIME: " + f"{total_run_time:.2f}" + " s" + "\n")
                f.write("RF Link Loss File (.xlsx): " + str(self.excel_file_var.get() or 'None') + "\n")
                f.write("RF Probe Loss File (.s2p): " + str(self.s2p_file_var.get() or 'None') + "\n")
                f.write("INITIAL PHOTOCURRENT: " + str(self.photo_currents[0]) + " (mA)" + "\n")
                f.write("STARTING WAVELENGTH FOR LASER 3: " + str(self.laser_3_var.get()) +
                        " (nm) : STARTING WAVELENGTH FOR LASER 4: " + f"{self.laser_4_wavelengths[0]:.3f}" +
                        " (nm) : DELAY: " + str(self.delay_var.get()) + " (s) " + "\n")
                f.write("DATE: " + time.strftime("%m/%d/%Y") + "\n")
                f.write("TIME: " + time.strftime("%H:%M:%S") + "\n")
                f.write("\n")
                f.write("F_BEAT(GHz)\tI_PD (mA)\tRaw RF POW (dBm)\tTotal RF Loss (dB)\tProbe RF Loss (dB)\tLink RF Loss (dB)\tCal RF POW (dBm)\tVOA P Actual (dBm)\n")
                for i in range(len(self.steps)):
                    f.write(f"{self.beat_freqs[i]:<10.2f}\t{self.photo_currents[i]:<10.4e}\t{self.powers[i]:<10.2f}\t"
                            f"{self.rf_loss[i]:<10.2f}\t{self.rf_probe_loss[i]:<10.2f}\t{self.rf_link_loss[i]:<10.2f}\t"
                            f"{self.calibrated_rf[i]:<10.2f}\t{self.p_actuals[i]:<10.3f}\n")
    
        except Exception as e:
            self.root.after(0, lambda: self.update_message_feed(f"Text save failed: {e}"))

        # 3) build & save Excel workbook
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Experiment Data"
            ws.append(["DEVICE NUMBER", self.device_num])
            ws.append(["COMMENTS", self.user_comment])
            ws.append(["KEITHLEY VOLTAGE", f"{self.keithley_voltage} V"])
            ws.append(["INITIAL PHOTOCURRENT", f"{self.photo_currents[0]} (mA)"])
            ws.append(["STARTING WAVELENGTH FOR LASER 3", f"{self.laser_3_var.get()} (nm)"])
            ws.append(["STARTING WAVELENGTH FOR LASER 4", f"{self.laser_4_wavelengths[0]:.3f} (nm)"])
            ws.append(["DELAY", f"{self.delay_var.get()} (s)"])
            ws.append(["FREQUENCY SWEEP RUN TIME", f"{sweep_run_time:.2f} s"])
            ws.append(["TOTAL RUN TIME", f"{total_run_time:.2f} s"])
            ws.append(["EXCEL LOSS FILE", self.excel_file_var.get() or 'None'])
            ws.append(["S2P LOSS FILE", self.s2p_file_var.get() or 'None'])
            ws.append(["DATE", time.strftime("%m/%d/%Y")])
            ws.append(["TIME", time.strftime("%H:%M:%S")])
            ws.append([])
            ws.append(["F_BEAT (GHz)", "I_PD (mA)", "Raw RF POW (dBm)",
                        "Total RF Loss (dB)", "RF Probe Loss (dB)",
                        "RF Link Loss (dB)", "Cal RF POW (dBm)", "VOA P Actual (dBm)"])
            for i in range(len(self.steps)):
                ws.append([
                    f"{self.beat_freqs[i]:.2f}",
                    f"{self.photo_currents[i]}",
                    f"{self.powers[i]:.2f}",
                    f"{self.rf_loss[i]:.2f}",
                    f"{self.rf_probe_loss[i]:.2f}",
                    f"{self.rf_link_loss[i]:.2f}",
                    f"{self.calibrated_rf[i]:.2f}",
                    f"{self.p_actuals[i]:.3f}"
                ])
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            wb.save(file_path.replace(".txt", ".xlsx"))
            self.update_message_feed(f"Excel data saved to {file_path.replace('.txt', '.xlsx')}")
    
        except Exception as e:
            self.root.after(0, lambda: self.update_message_feed(f"Excel save failed: {e}"))

        # 4) all done!
        self.root.after(0, lambda: self.update_message_feed(
            f"Data & plot saved to {file_path} and {plot_file_path}"
        ))
 
    def update_plots(self):
        """
        Update the Matplotlib plots with the latest data.
        This method is repeatedly called using Tkinter's after() method.
        """
        if self.data_ready_event.is_set():
            self.line1.set_data(self.steps,   self.beat_freqs)
            self.markers1.set_data(self.steps,   self.beat_freqs)
            self.line2.set_data(self.steps,   self.laser_4_wavelengths)
            self.markers2.set_data(self.steps,   self.laser_4_wavelengths)
            self.line3.set_data(self.beat_freqs, self.powers)
            self.markers3.set_data(self.beat_freqs, self.powers)
            self.line4.set_data(self.beat_freqs, self.photo_currents)
            self.markers4.set_data(self.beat_freqs, self.photo_currents)
            if not self.looping:
                self.line5.set_data(self.beat_freqs, self.calibrated_rf)
                self.markers5.set_data(self.beat_freqs, self.calibrated_rf)
            for ax in (self.ax1, self.ax2, self.ax3, self.ax4, self.ax5):
                ax.relim(); ax.autoscale_view()
            self.canvas.draw()
            self.data_ready_event.clear()

        # unconditionally keep the GUI refresher running
        self.root.after(100, self.update_plots)

    def on_stop(self):
        """
        Handle the "Stop" button press.
        Ask the user for confirmation and set the stop_event to end data collection.
        """
        if messagebox.askyesno("Confirm Stop", "Are you sure you want to stop the data collection?"):
            self.stop_event.set()
            self.update_message_feed("Data collection will be stopped.")

    def reset_program(self):
        """
        Reset only the measurement data arrays, leaving file paths and comments intact.
        Also reinitialize instruments.

        """

        self.looping = False
        # Clear only measurement data containers.
        self.steps = []
        self.beat_freqs = []
        self.laser_4_wavelengths = []
        self.beat_freq_and_power = []
        self.calibrated_rf = []
        self.photo_currents = []
        self.rf_loss = []
        self.rf_probe_loss = []
        self.rf_link_loss = []
        self.powers = []
        self.p_actuals = []

        # Optionally, remove any text annotations you previously added.
        texts_to_remove = [txt for txt in self.fig.texts if txt != self.fig._suptitle]
        for txt in texts_to_remove:
            txt.remove()

        # Reset plot data
        self.line1.set_data([], [])
        self.markers1.set_data([], [])
        self.line2.set_data([], [])
        self.markers2.set_data([], [])
        self.line3.set_data([], [])
        self.markers3.set_data([], [])
        self.line4.set_data([], [])
        self.markers4.set_data([], [])
        self.line5.set_data([], [])
        self.markers5.set_data([], [])
        for ax in [self.ax1, self.ax2, self.ax3, self.ax4, self.ax5]:
            ax.relim()
            ax.autoscale_view()
        self.canvas.draw()
        self.root.after(100, self.update_plots)

        # Clear stop and data_ready events.
        self.stop_event.clear()
        self.data_ready_event.clear()
        self.update_message_feed("Program reset and ready to start again.")

        # Reinitialize instruments.
        self.close_instruments()
        self.open_instruments()



    def on_cancel(self):
        """
        Handle the "RESET" button.
        Confirm with the user before resetting the program.
        """
        if messagebox.askyesno("Confirm Exit", "Are you sure you want to reset the program?"):
            self.stop_event.set()  # Signal the thread to stop
            # Wait for the measurement thread to finish (wait up to 5 seconds)
            if hasattr(self, 'measurement_thread'):
                timeout = 5  # total seconds to wait
                interval = 0.5  # check interval
                waited = 0
                while self.measurement_thread.is_alive() and waited < timeout:
                    time.sleep(interval)
                    waited += interval
            self.reset_program()


    def on_closing(self):
        """
        Handle window closing.
        Confirm with the user before quitting and cleanly exit.
        """
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.stop_event.set()
            self.root.destroy()
            sys.exit(0)

    def run(self):
        """Start the Tkinter main loop."""
        self.root.mainloop()


if __name__ == '__main__':
    app = MeasurementApp()
    app.run()
    app.close_instruments()
