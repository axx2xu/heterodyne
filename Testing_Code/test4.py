import time
import math
import pyvisa
import sys
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import Workbook
import threading
import msvcrt
import re
from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter import messagebox
import tkinter.filedialog as fd



################################################################################################################################################################################
#                                   **** THIS PROGRAM RUNS AUTOMATIC CALIBRATION AND LOOPING FOR BEAT FREQUENCY MEASUREMENTS ****
#                             **** THE PROGRAM RECORDS THE BEAT FREQUENCY, PHOTOCURRENT, RAW RF POWER, AND CALIBRATED RF POWER ****
#                        **** THE PROGRAM ALSO PLOTS THE BEAT FREQUENCY, LASER 4 WAVELENGTH, RAW RF POWER, AND CALIBRATED RF POWER **** 
#                                   **** THE DATA IS THEN OUTPUT TO A .txt FILE IF THE USER SAYS YES WHEN PROMPTED ****
#
################################################################################################################################################################################



############################################################################################################################################################
####                                            GPIB ADDRESS THAT NEED TO BE MANUALLY UPDATED BY THE USER                                               ###
############################################################################################################################################################

# Initialize the VISA resource manager
rm = pyvisa.ResourceManager()

# List all connected VISA devices (Optional: To verify connections)
print("Connected devices:", rm.list_resources())

# Create a VISA adapter for the ECL laser and wavelength meter
# *** The GPIB should always be the same, so these should not need to be changed ***

ecl_adapter = rm.open_resource('GPIB0::10::INSTR')  # Update with your actual GPIB address
wavelength_meter = rm.open_resource('GPIB0::20::INSTR')  # Update with your actual GPIB address
spectrum_analyzer = rm.open_resource('GPIB0::18::INSTR')  # Update with your actual GPIB address
keithley = rm.open_resource('GPIB0::24::INSTR')  # Update with your actual GPIB address
RS_power_sensor = rm.open_resource('RSNRP::0x00a8::100940::INSTR') # Update with your actual VISA address for the RS NRP-Z58 sensor
voa = rm.open_resource('GPIB0::26::INSTR')  # Update with your actual GPIB address

############################################################################################################################################################
####                                                  METHODS FOR REPEATED MEASUREMENT OR INSTRUMENT USE                                                 ###
############################################################################################################################################################

# Disconnects from instruments and exits the program using (Ctrl + C)
def exit_program():
    """Exit program and close the connection."""
    ecl_adapter.close()
    wavelength_meter.close()
    spectrum_analyzer.close()
    keithley.close()
    voa.close()
    RS_power_sensor.close()
    print("Connection closed.")
    sys.exit(0)

# Defined function to find the peak frequency using the peak search function on the ESA
def measure_peak_frequency(spectrum_analyzer):
    """Measure the peak frequency using the spectrum analyzer."""
    try:
        """Perform peak search and return the peak frequency within the current span."""
        spectrum_analyzer.write('MKPK HI')
        time.sleep(0.1)  # Allow time for the peak search to complete
        peak_freq_1 = spectrum_analyzer.query('MKF?')
        time.sleep(0.1)
        spectrum_analyzer.write('MKPK HI')
        time.sleep(0.1)  # Allow time for the peak search to complete
        peak_freq_2 = spectrum_analyzer.query('MKF?')
        time.sleep(0.1)
        spectrum_analyzer.write('MKPK HI')
        time.sleep(0.1)  # Allow time for the peak search to complete
        peak_freq_3 = spectrum_analyzer.query('MKF?')

        peak_freq = min(peak_freq_1, peak_freq_2, peak_freq_3)

        return float(peak_freq) / 1e9 # Convert Hz to GHz
    except Exception as e:
        print("Error measuring peak frequency with spectrum analyzer:", e)
        return None

# Defined function to measure the beat frequency using the wavelength meter
def measure_wavelength_beat(wavelength_meter):
    """Measure the beat frequency using the wavelength meter."""
    try:
        wavelength_meter.write(":INIT:IMM") # Initiate a single measurement
        time.sleep(1.5)  # Increased wait time for the measurement to complete
        freq_data = wavelength_meter.query(":CALC3:DATA? FREQuency").strip().split(',') # Query to get the frequency, strip to only get the first two values (corresponding to laser 3 and 4)
        freqs = [float(freq) for freq in freq_data]
        if len(freqs) == 2:
            return abs(freqs[0] - freqs[1]) / 1e9  # Calculate the difference in GHz
    except Exception as e:
        print("Error measuring beat frequency with wavelength meter:", e)
        return None

# Defined functions to set the laser wavelength and power
def set_laser_wavelength(ecl_adapter, channel, wavelength):
    """Set the laser wavelength."""
    print(f"Setting laser {channel} wavelength to {wavelength:.3f} nm...")
    ecl_adapter.write(f"CH{channel}:L={wavelength:.3f}")

# Function to read Excel data using openpyxl
def read_excel_data(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    # Assuming data is in the first two columns
    frequency = []
    loss = []
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        frequency.append(row[0])
        loss.append(row[1])

    return np.array(frequency), np.array(loss)

def read_s2p_file(filepath):
    frequencies = []
    s12 = []
    s21 = []
    
    with open(filepath, 'r') as file:
        lines = file.readlines()
    
    freq_unit = 'hz'
    data_format = 'db'
    
    for line in lines:
        if line.startswith('#'):
            parts = line.split()
            for part in parts:
                if part.lower() in ['hz', 'khz', 'mhz', 'ghz']:
                    freq_unit = part.lower()
                elif part.lower() in ['ri', 'db', 'ma']:
                    data_format = part.upper()
            continue
        
        if not line.startswith('!') and not line.startswith('#'):  # Ignore comment and header lines
            values = re.split(r'\s+', line.strip())
            if len(values) >= 9:  # Ensure there are enough values in the line
                frequencies.append(float(values[0]))
                s21.append(float(values[3]))  # S21 in dB
                s12.append(float(values[5]))  # S12 in dB
    
    frequencies = np.array(frequencies)
    s_avg = (np.array(s12) + np.array(s21)) / 2

    # Convert frequencies to GHz if needed
    if freq_unit == 'khz':
        frequencies = frequencies / 1e6
    elif freq_unit == 'mhz':
        frequencies = frequencies / 1e3
    elif freq_unit == 'hz':
        frequencies = frequencies / 1e9
    
    return frequencies, s_avg

def custom_linear_interpolation(x, y, x_new):
    """Perform linear interpolation on given data points."""
    x = np.asarray(x)
    y = np.asarray(y)
    x_new = np.asarray(x_new)
    y_new = np.zeros_like(x_new)

    for i, xi in enumerate(x_new):
        if xi <= x[0]:
            y_new[i] = y[0]
        elif xi >= x[-1]:
            y_new[i] = y[-1]
        else:
            # Find the interval [x_k, x_k+1] where x_k <= xi < x_k+1
            k = np.searchsorted(x, xi) - 1
            x_k = x[k]
            x_k1 = x[k + 1]
            y_k = y[k]
            y_k1 = y[k + 1]

            # Linear interpolation formula
            y_new[i] = y_k + (y_k1 - y_k) * (xi - x_k) / (x_k1 - x_k)

    return y_new

flag = True
exit_event = threading.Event()

def exit_loop():
    global flag
    print('Press any key to stop the loop')
    while not exit_event.is_set():
        if msvcrt.kbhit():  # Check for key press
            msvcrt.getch()  # Clear the key press
            print('Stopping the loop...')
            flag = False
            break
        time.sleep(0.1)  # Check for key press every 100ms


# Set timeout to 5 seconds (5000 milliseconds)
ecl_adapter.timeout = 5000
wavelength_meter.timeout = 5000
spectrum_analyzer.timeout = 5000
keithley.timeout = 5000
voa.timeout = 5000
RS_power_sensor.timeout = 5000

# Initialize the Tkinter root widget
root = Tk()
root.withdraw()  # Hide the root window

try:
    # Set the Keithley to local mode
    keithley.write(":SYSTem:LOCal")

    print("Please enter the parameters for the measurement:")

    # Value for laser 3 input wavelength check
    while True:
        laser_3_WL = float(input("Enter the starting WL for laser 3 (nm): "))
        if 1540 <= laser_3_WL <= 1640:
            break
        else:
            print("Invalid input. Please enter a wavelength between 1540 and 1640 nm.")

    # Value for laser 4 input wavelength check
    while True:
        laser_4_WL = float(input("Enter the starting WL for laser 4 (nm) (recommend 2 nm below Laser 3): "))
        if 1540 <= laser_4_WL <= 1640:
            break
        else:
            print("Invalid input. Please enter a wavelength between 1540 and 1640 nm.")

    # Start beat frequency 
    start_freq = float(input("Enter the starting beat frequency (GHz): "))

    # End beat frequency
    end_freq = float(input("Enter the final beat frequency (GHz): "))

    # Value for the final frequency stepped to (i.e. 1 GHz step size with final frequency of 100 GHz will have 100 steps)
    # final_freq = float(input("Enter the final frequency (GHz): "))
    num_steps = int(input("Enter the number of steps you would like the program to take: "))

    # Enter a delay time for the lasers to stabilize
    try:
        delay = int(input("Enter the delay time for the lasers to stabilize (s), Recommended 3+ seconds: "))
    except ValueError:
        print("Invalid input, using default input of 3 seconds")
        delay = 3

    # Ask the user if they want to use an s2p file for calibration
    s2p_input = input("Would you like to use an s2p file for calibration? (Yes (Y) /No (N)): ").strip().upper()
    while s2p_input not in ['YES', 'NO', 'Y', 'N']:  # Loop until get a valid input
        s2p_input = input("Invalid input. Would you like to use an s2p file for calibration? (Yes (Y) /No (N)): ").strip().upper()

    if s2p_input in ['YES', 'Y']:
        root.lift()  # Bring the root window to the front
        root.focus_force()  # Focus the root window
        root.attributes('-topmost', True)  # Make the root window stay on top
        time.sleep(0.1)  # Short delay to ensure the window manager processes the change
        s2p_filename = fd.askopenfilename(
            title=  "Select s2p File",
            filetypes=[("s2p", "*.s2p")]
        )
        root.attributes('-topmost', False)  # Disable the topmost attribute
        if s2p_filename:
            print(f"Selected s2p file: {s2p_filename}")
        else:
            print("No s2p file selected.")
    

    # Ask the user if they want to use an excel file for calibration
    excel_input = input("Would you like to use an excel file for calibration? (Yes (Y) /No (N)): ").strip().upper()
    while excel_input not in ['YES', 'NO', 'Y', 'N']:  # Loop until get a valid input
        excel_input = input("Invalid input. Would you like to use an excel file for calibration? (Yes (Y) /No (N)): ").strip().upper()

    if excel_input in ['YES', 'Y']:
        root.lift()  # Bring the root window to the front
        root.focus_force()  # Focus the root window
        root.attributes('-topmost', True)  # Make the root window stay on top
        time.sleep(0.1)  # Short delay to ensure the window manager processes the change
        excel_filename = fd.askopenfilename(
            title=  "Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        root.attributes('-topmost', False)  # Disable the topmost attribute
        if excel_filename:
            print(f"Selected Excel file: {excel_filename}")
        else:
            print("No Excel file selected.")


    # Ask the user if they want to perform automatic calibration
    calibration_input = input("Would you like to run automatic start frequency search? (Yes (Y) /No (N)): ").strip().upper()
    while calibration_input not in ['YES', 'NO', 'Y', 'N']:  # Loop until get a valid input
        calibration_input = input("Invalid input. Would you like to run automatic calibration? (Yes (Y) /No (N)): ").strip().upper()

    if calibration_input in ['YES', 'Y']:

        # User input for frequency threshold
        user_input = input("Use default threshold of 1GHz (Enter D) or custom threshold in GHz (Enter C) (NOTE: Values under 1GHz are not guaranteed to work): ").strip().upper()
        if user_input == 'D':
            freq_threshold = 1
        elif user_input == 'C':
            freq_threshold = float(input("Enter frequency threshold (GHz): "))
        else:
            print("Invalid input. Using default frequency threshold of 1 GHz.")
            freq_threshold = 1

    

    # Lists to store step number, beat frequency, and laser 4 wavelength
    steps = []
    beat_freqs = []
    laser_4_wavelengths = []
    beat_freq_and_power = []

    # Set the laser wavelengths and power
    set_laser_wavelength(ecl_adapter, 3, laser_3_WL)
    set_laser_wavelength(ecl_adapter, 4, laser_4_WL)

    # Wait for the lasers to stabilize
    print("Waiting for the lasers to stabilize...")
    time.sleep(10)

    c = 299792458  # Speed of light in m/s

    # Set the reference frequency to laser 3
    laser_3_freq = c / (laser_3_WL * 1e-9)  # Convert nm to meters and calculate frequency
    wavelength_meter.write(f":CALC3:DELTA:REF:FREQ {laser_3_freq}")
    time.sleep(1)  # Small delay to ensure command is processed

    # Store the initial wavelength of Laser 4
    initial_laser_4_WL = laser_4_WL

    # Additional variables to track consecutive increases
    consecutive_increases = 0
    max_consecutive_increases = 3  # Number of consecutive increases to trigger recalibration

    # Store the initial wavelength of Laser 4
    initial_laser_4_WL = laser_4_WL
    below_threshold = False

    # Measure beat frequency using wavelength meter and ESA
    wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
    esa_beat_freq = measure_peak_frequency(spectrum_analyzer)

    # If wl_meter_beat_freq is None, it means the wavelength meter failed to measure, so we use the ESA value
    if wl_meter_beat_freq is None:
        wl_meter_beat_freq = esa_beat_freq

    current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq

    # Initialize last_beat_freq
    last_beat_freq = None

    if calibration_input in ['YES', 'Y']:
        ############################################################################################################################################################
        ####                                                   CALIBRATION TO GET BEAT FREQUENCY WITHIN 1 GHz                                                    ###
        ############################################################################################################################################################
        # Loop through until the starting frequency is within the given threshold
        print("RUNNING CALIBRATION LOOP...")
        while current_freq >= 0.5:

            # Calculate the new wavelength for laser 4
            laser_4_freq = c / (laser_4_WL * 1e-9)  # Calculate current frequency of laser 4

            # Measure beat frequency using wavelength meter and ESA
            wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
            esa_beat_freq = measure_peak_frequency(spectrum_analyzer)

            # If wl_meter_beat_freq is None, it means the wavelength meter failed to measure, so we use the ESA value
            if wl_meter_beat_freq is None:
                wl_meter_beat_freq = esa_beat_freq

            if wl_meter_beat_freq is None or esa_beat_freq is None:
                continue

            current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
            #print(f"Current Beat Frequency: {current_freq} GHz")

            if last_beat_freq is not None and current_freq > last_beat_freq:
                if current_freq >= 1:
                    consecutive_increases += 1
                    if consecutive_increases >= max_consecutive_increases:
                        print("Calibration overshot the threshold, restarting calibration...")
                        # Reset calibration variables
                        laser_4_WL = laser_3_WL - 2 # Reset to initial wavelength - 1 nm
                        set_laser_wavelength(ecl_adapter, 4, laser_4_WL)  # Apply the reset wavelength to Laser 4
                        consecutive_increases = 0
                        last_beat_freq = None  # Reset last beat frequency
                        time.sleep(15)  # Give some time before restarting the loop

                        # Measure beat frequency using wavelength meter and ESA
                        wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
                        esa_beat_freq = measure_peak_frequency(spectrum_analyzer)

                        # If wl_meter_beat_freq is None, it means the wavelength meter failed to measure, so we use the ESA value
                        if wl_meter_beat_freq is None:
                            wl_meter_beat_freq = esa_beat_freq

                        if wl_meter_beat_freq is None or esa_beat_freq is None:
                            continue

                        current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
                        continue
                    
                elif current_freq < 1:
                    break #Near 0 beat frequency
            else:
                consecutive_increases = 0

            # Update last_beat_freq at the end of the loop
            

            if wl_meter_beat_freq >= 50:
                print(f"Beat Frequency (Wavelength Meter): {wl_meter_beat_freq} GHz")
                laser_4_new_freq = laser_4_freq - (wl_meter_beat_freq * 0.67 * 1e9) # Update the frequency by 2/3 of the beat frequency

                laser_4_WL = (c / laser_4_new_freq) * 1e9  # Set the new wavelength

                # Check if the new wavelength is within the bounds of the ECL laser and different from the current wavelength
                if 1540 < laser_4_WL < 1660:
                    set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                else:
                    print(f"New wavelength for laser 4 is out of bounds or unchanged: {laser_4_WL:.3f} nm")
                    exit_program()

            elif esa_beat_freq < 50 and wl_meter_beat_freq < 50:
                print(f"Beat Frequency (ESA): {round(esa_beat_freq,1)} GHz")

                if esa_beat_freq > 3:
                    if last_beat_freq is not None and last_beat_freq < 1:
                        laser_4_new_freq = laser_4_freq - (0.2 * 1e9)
                    else:
                        laser_4_new_freq = laser_4_freq - (esa_beat_freq * 0.67 * 1e9) # Update the frequency by 2/3 of the beat frequency for all beat freq greater than 3 GHz
                elif 1.5 < esa_beat_freq <= 3:
                    laser_4_new_freq = laser_4_freq - (0.5 * 1e9)
                elif 1 <= esa_beat_freq <= 1.5:
                    laser_4_new_freq = laser_4_freq - (0.2 * 1e9)
                # elif esa_beat_freq < 1:  # the value is within the threshold so break out of the loop as to not update the wavelength again unnecessarily
                #     break
                elif esa_beat_freq < 1:
                    laser_4_new_freq = laser_4_freq - (0.1 * 1e9)

                laser_4_WL = (c / laser_4_new_freq) * 1e9
                
                # Check if the new wavelength is within the bounds of the ECL laser and different from the current wavelength
                if 1540 < laser_4_WL < 1660:
                    set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                else:
                    print(f"New wavelength for laser 4 is out of bounds or unchanged: {laser_4_WL:.3f} nm")
                    exit_program()
            last_beat_freq = current_freq

            time.sleep(3)

        # Once within threshold of 0, update the step to attempt to jump over potential ESA measurement issues
        laser_4_freq = c / (laser_4_WL * 1e-9)  # Calculate current frequency of laser 4
        laser_4_new_freq = laser_4_freq - (1 * 1e9) # Update the frequency by half the difference between the current and starting frequency
        laser_4_WL = (c / laser_4_new_freq) * 1e9 # Set the new wavelength
        set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
        time.sleep(3) # Small delay before next iteration

        #Measure beat frequency using wavelength meter and ESA
        wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
        esa_beat_freq = measure_peak_frequency(spectrum_analyzer)

        # If wl_meter_beat_freq is None, it means the wavelength meter failed to measure, so we use the ESA value
        if wl_meter_beat_freq is None:
            wl_meter_beat_freq = esa_beat_freq

        current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq



        ############################################################################################################################################################
        ####                                                   SECOND LOOP TO GET UP TO STARTING FREQUENCY IF NEEDED                                             ###
        ############################################################################################################################################################

        update_laser = True

        while abs(current_freq - start_freq) > freq_threshold:

            #Measure beat frequency using wavelength meter and ESA
            wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
            esa_beat_freq = measure_peak_frequency(spectrum_analyzer)

            # If wl_meter_beat_freq is None, it means the wavelength meter failed to measure, so we use the ESA value
            if wl_meter_beat_freq is None:
                wl_meter_beat_freq = esa_beat_freq

            if wl_meter_beat_freq is None or esa_beat_freq is None:
                continue

            current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
            print(f"Current Beat Frequency: {round(current_freq,1)} GHz")

            laser_4_freq = c / (laser_4_WL * 1e-9)  # Calculate current frequency of laser 4
            laser_4_new_freq = laser_4_freq - ((abs(start_freq - current_freq) * 1e9))/2 # Update the frequency by half the difference between the current and starting frequency
            laser_4_WL = (c / laser_4_new_freq) * 1e9 # Set the new wavelength


            if(abs(current_freq - start_freq)) <= freq_threshold: # If the current frequency is within the threshold, don't update laser 4 again
                update_laser = False

            if(update_laser):
                # Check if the new wavelength is within the bounds of the ECL laser and different from the current wavelength
                if 1540 < laser_4_WL < 1660:
                    set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                else:
                    print(f"New wavelength for laser 4 is out of bounds or unchanged: {laser_4_WL:.3f} nm")
                    exit_program()

            time.sleep(3) # Small delay before next iteration
            last_beat_freq = current_freq # Update last frequency
        
        time.sleep(5)
        #Measure beat frequency using wavelength meter and ESA
        wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
        esa_beat_freq = measure_peak_frequency(spectrum_analyzer)

        # If wl_meter_beat_freq is None, it means the wavelength meter failed to measure, so we use the ESA value
        if wl_meter_beat_freq is None:
            wl_meter_beat_freq = esa_beat_freq

        current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
        print(f"Current Beat Frequency: {round(current_freq,1)} GHz")
        
    laser_4_cal = round(laser_4_WL,3) # Store the calibrated wavelength for laser 4 to use in txt file header

    
    ############################################################################################################################################################
    ####                                             LOOP THROUGH STEPS, UPDATE WAVELENGTHS, TAKE MEASUREMENTS                                               ###
    ############################################################################################################################################################

    print("BEGINNING MEASUREMENT LOOP...")

    time_start = time.time()  # Start time of the loop

    last_beat_freq = current_freq
    # Initialize last beat frequency to the current frequency

    plt.ion()  # Turn on interactive mode for live plotting

    fig, axes = plt.subplots(2, 2, figsize=(8, 8))
    fig.suptitle(f"Center Wavelength: {laser_3_WL:.2f} nm, Delay: {delay} s", fontsize=16)
    ax1, ax3, ax4, ax5 = axes[0, 0], axes[0, 1], axes[1, 0], axes[1, 1]

    ax2 = ax1.twinx()  # Create a twin y-axis for the first subplot

    # Setup first subplot (ax1)
    color = 'tab:blue'
    ax1.set_xlabel('Step Number')
    ax1.set_ylabel('Beat Frequency (GHz)', color=color)
    ax1.set_title('Beat Frequency vs Step Number')
    line1, = ax1.plot([], [], marker='o', linestyle='-', color=color)
    ax1.tick_params(axis='y', labelcolor=color)
    ax1.grid(True)

    color = 'tab:red'
    ax2.set_ylabel('Laser 4 Wavelength (nm)', color=color)
    line2, = ax2.plot([], [], marker='x', linestyle='--', color=color)
    ax2.tick_params(axis='y', labelcolor=color)
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:.3f}'.rstrip('0').rstrip('.')))


    # Setup second subplot (ax3)
    color = 'tab:blue'
    ax3.set_xlabel('Beat Frequency (GHz)')
    ax3.set_ylabel('Raw RF Power (dBm)', color=color)
    ax3.set_title('Raw RF Power vs Beat Frequency')
    line3, = ax3.plot([], [], marker='o', linestyle='-', color=color)
    ax3.tick_params(axis='y', labelcolor=color)
    ax3.grid(True)

    # Setup third subplot (ax4)
    color = 'tab:blue'
    ax4.set_xlabel('Beat Frequency (GHz)')
    ax4.set_ylabel('Photocurrent (mA)', color=color)
    ax4.set_title('Measured Photocurrent vs Beat Frequency')
    line4, = ax4.plot([], [], marker='o', linestyle='-', color=color)
    ax4.tick_params(axis='y', labelcolor=color)
    ax4.grid(True)

    # Setup fourth subplot (ax5)
    color = 'tab:blue'
    ax5.set_xlabel('Beat Frequency (GHz)')
    ax5.set_ylabel('Calibrated RF Power (dBm)', color=color)
    ax5.set_title('Calibrated RF Power vs Beat Frequency')
    line5, = ax5.plot([], [], marker='o', linestyle='-', color=color)
    ax5.tick_params(axis='y', labelcolor=color)
    ax5.grid(True)

    fig.tight_layout()

    # Maximize the window for different backends
    manager = plt.get_current_fig_manager()
    try:
        # TkAgg backend
        manager.window.state('zoomed')
    except AttributeError:
        try:
            # Qt5Agg backend
            manager.window.showMaximized()
        except AttributeError:
            try:
                # WxAgg backend
                manager.window.Maximize(True)
            except AttributeError:
                pass  # If none of these work, just continue
    
    # Calculate step size
    laser_4_step = (end_freq - start_freq) / num_steps  # Calculate the step size for laser 4 wavelength

    # Get initial photocurrent for text output
    # MEASURE CURRENT FROM KEITHLEY
    response = keithley.query(":MEASure:CURRent?")
    initial_current_values = response.split(',')

    if len(initial_current_values) > 1:
        initial_current = float(initial_current_values[1]) * 1000  # Convert to mA
        initial_current = round(initial_current,3)  # Format for display

    # Set the Keithley back to local mode
    keithley.write(":SYSTem:LOCal")

    n = threading.Thread(target=exit_loop) # Start thread to be able to cancel the loop by keystroke
    n.start() # Start the thread that will look for a keystroke in order to stop the measurement loop early
    try:
        for step in range(num_steps):
            if not flag:
                break
            beat_freq = measure_peak_frequency(spectrum_analyzer) if last_beat_freq < 45 else measure_wavelength_beat(wavelength_meter)
            if beat_freq is None:
                continue

            print(f"Step Number: {step + 1}")

            if(step < 5 and start_freq < 5 and beat_freq > 10):
                laser_4_freq = c / (laser_4_WL * 1e-9)
                laser_4_new_freq = laser_4_freq - (0.3 * 1e9) # Within the threshold where ESA doesn't work, so just decrease by 0.1 GHz
                laser_4_WL = (c / laser_4_new_freq) * 1e9
                set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                time.sleep(delay) # Delay for the lasers to stabilize

                # Re-measure the beat frequency
                beat_freq = measure_peak_frequency(spectrum_analyzer) if last_beat_freq < 45 else measure_wavelength_beat(wavelength_meter)
                if beat_freq is None:
                    continue

            print(f"Beat Frequency: {round(beat_freq,1)} GHz")

            # MEASURE CURRENT FROM KEITHLEY
            response = keithley.query(":MEASure:CURRent?")
            current_values = response.split(',')

            if len(current_values) > 1:
                current = float(current_values[1]) * 1000  # Convert to mA
                current = round(current,3)  # Format for display
                print(f"Measured Photocurrent: {current} (mA)")

            # Set the Keithley back to local mode
            keithley.write(":SYSTem:LOCal")

            max_attempts = 3
            attempts = 0
            success = False

            # Measure the VOA P actual
            # Trigger a measurement
            p_actual = voa.query('READ:POW?')
            p_actual = float(p_actual) # convert to float so it can be rounded
            p_actual = round(p_actual,3) # round 3 decimal places
            voa.write('SYST:LOC')  # Put the VOA back into local mode so the attenuation can be adjusted

            # Added loop to retry the measurement if it fails, have encountered random errors with the power sensor
            while attempts < max_attempts and not success:
                try:
                    # WRITE TO THE R&S POWER SENSOR
                    RS_power_sensor.write('INIT:CONT OFF')
                    RS_power_sensor.write('SENS:FUNC "POW:AVG"')
                    RS_power_sensor.write(f'SENS:FREQ {beat_freq}e9')  # Update with new beat frequency
                    RS_power_sensor.write('SENS:AVER:COUN:AUTO ON')
                    RS_power_sensor.write('SENS:AVER:STAT ON')
                    RS_power_sensor.write('SENS:AVER:TCON REP')
                    RS_power_sensor.write('SENS:POW:AVG:APER 5e-3')

                    rf_outputs = []

                    for i in range(5):
                        RS_power_sensor.write('INIT:IMM')
                        time.sleep(0.1) # Wait for the measurement to complete
                        output = RS_power_sensor.query('TRIG:IMM')
                        output = output.split(',')[0]  # split the output to only get first value (power in Watts)
                        rf_outputs.append(float(output))
                        time.sleep(0.1) # Wait for the measurement to complete
                    
                    output = sum(rf_outputs) / len(rf_outputs)  # Calculate the average of the 5 measurements
                    output_dbm = math.log10(output) * 10 + 30  # convert from watts to dBm
                    
                    output_dbm = round(output_dbm, 2)  # round to 2 decimal places
                    beat_freq = round(beat_freq, 2)  # Round beat frequency to 1 decimal place
                    
                    beat_freq_and_power.append((beat_freq, output_dbm, current, p_actual))
                    success = True  # Measurement succeeded
                except ValueError as e:
                    print(f"Error processing measurement at step {step + 1}, attempt {attempts + 1}: {e}")
                    attempts += 1
                    time.sleep(1)  # Wait a bit before retrying

            if not success:
                print(f"Measurement failed after {max_attempts} attempts at step {step + 1}")
                continue


            last_beat_freq = beat_freq
            steps.append(step + 1)
            beat_freqs.append(beat_freq)
            laser_4_wavelengths.append(laser_4_WL)

            laser_4_freq = c / (laser_4_WL * 1e-9)
            laser_4_new_freq = laser_4_freq - (laser_4_step * 1e9)
            laser_4_WL = (c / laser_4_new_freq) * 1e9
            set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
            

            # Update the plots with the new data
            line1.set_data(steps, beat_freqs)
            line2.set_data(steps, laser_4_wavelengths)
            line3.set_data(beat_freqs, [x[1] for x in beat_freq_and_power])
            line4.set_data(beat_freqs, [x[2] for x in beat_freq_and_power])
            ax1.relim()
            ax2.relim()
            ax3.relim()
            ax4.relim()
            ax1.autoscale_view()
            ax2.autoscale_view()
            ax3.autoscale_view()
            ax4.autoscale_view()
            plt.draw()
            plt.pause(0.1)

            last_beat_freq = beat_freq

            time.sleep(delay)  # Delay for the lasers to stabilize based on user input
    finally:
        # Signal the exit loop thread to stop
        exit_event.set() # Set the event to stop the thread from looking for a key stroke
        n.join()  # Wait for the thread to finish
        print("Loop stopped.")

    plt.ioff()  # Turn off interactive mode

    time_end = time.time()  # End time of the loop
    run_time = time_end - time_start  # Calculate the total run time
    
    # Sort the data by beat frequency
    beat_freq_and_power.sort(key=lambda x: x[0])
    beat_freqs_pow, powers, photo_currents, p_actuals = zip(*beat_freq_and_power)

    ############################################################################################################################################################
    #                                   Calculate Calibrated RF from network analyzer file
    ############################################################################################################################################################

    # Replace the interpolation part with custom_linear_interpolation

    calibrated_rf = np.array(powers)  # Initialize the calibrated RF power with the raw RF power
    rf_loss = np.zeros_like(calibrated_rf)   # Initialize the list to store the RF loss values

    try:
        # Read the .s2p file and extract data
        frequencies, s_avg = read_s2p_file(s2p_filename)

        # Interpolate the data
        interpolated_loss = custom_linear_interpolation(frequencies, s_avg, beat_freqs_pow)

        # Make into np array
        interpolated_loss = np.array(interpolated_loss)

        # Now update the calibrated_rf calculation to include the new interpolated loss values from s2p file
        rf_loss += np.abs(interpolated_loss)

        calibrated_rf += np.abs(interpolated_loss.real)
        calibrated_rf = np.round(calibrated_rf, 2)  # Round the data
    except Exception as e:
        print("No network analyzer file found or error in processing:", e)

    try:
        # Gather excel data for RF probe loss
        probe_loss_frequency, probe_loss = read_excel_data(excel_filename)
        interpolated_probe_loss = custom_linear_interpolation(probe_loss_frequency, probe_loss, beat_freqs_pow)

        interpolated_probe_loss = np.array(interpolated_probe_loss)

        # Now update the calibrated_rf calculation to include the new interpolated loss values from excel file
        rf_loss += np.abs(interpolated_probe_loss)

        calibrated_rf += np.abs(interpolated_probe_loss)
        calibrated_rf = np.round(calibrated_rf, 2)  # Round the data
    except Exception as e:
        print("No excel file found or error in processing:", e)

    ############################################################################################################################################################
    ####                                              PROMPTED USER INPUTS FOR OUTPUTTING DATA IN .TXT FILE                                                  ###
    ############################################################################################################################################################

    output_input = input("Would you like to output the data to a .txt file? (Yes/No): ").strip().upper()
    while output_input not in ['YES', 'NO', 'Y', 'N']:  # Loop until get a valid input
        output_input = input("Invalid input. Would you like to output the data to a .txt file? (Yes (Y) / No (N)): ").strip().upper()

    if output_input in ['YES', 'Y']:
        # Get user input for file name and path
        root.lift()  # Bring the root window to the front
        root.focus_force()  # Focus the root window
        root.attributes('-topmost', True)  # Make the root window stay on top
        time.sleep(0.1)  # Short delay to ensure the window manager processes the change
        file_path = fd.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        root.attributes('-topmost', False)  # Disable the topmost attribute
        if file_path:
            device_num = input("Enter the device number: ").strip()
            comment = input("Enter any trial comments: ").strip().upper()
            keithley_voltage = keithley.query(":SOUR:VOLT:LEV:IMM:AMPL?").strip()  # Get the keithley voltage from the keithley
            keithley_voltage = f"{float(keithley_voltage):.3e}"  # Format the keithley voltage for display
            keithley.write(":SYSTem:LOCal")  # Set the keithley back to local mode

            with open(file_path, 'w') as f:
                f.write("DEVICE NUMBER: " + str(device_num) + "\n")
                f.write("COMMENTS: " + comment + "\n")
                f.write("KEITHLEY VOLTAGE: " + str(keithley_voltage) + " V" + "\n")
                f.write("INITIAL PHOTOCURRENT: " + str(initial_current) + " (mA)" + "\n")
                f.write("STARTING WAVELENGTH FOR LASER 3: " + str(laser_3_WL) + " (nm) :" + " STARTING WAVELENGTH FOR LASER 4: " + str(laser_4_cal) + " (nm) :" + " DELAY: " + str(delay) + " (s) " + "\n")
                f.write("DATE: " + time.strftime("%m/%d/%Y") + "\n")
                f.write("TIME: " + time.strftime("%H:%M:%S") + "\n")
                f.write("RUN TIME: " + str(run_time) + " (s)" + "\n")
                f.write("\n")
                f.write("F_BEAT(GHz)\tPHOTOCURRENT (mA)\tRaw RF POW (dBm)\tRF Loss (dB)\t\tCal RF POW (dBm)\tVOA P Actual (dBm)\n")
                for i in range(len(steps)):
                    f.write(f"{beat_freqs_pow[i]:<10.2f}\t{float(photo_currents[i]):<10.4e}\t\t{powers[i]:<10.2f}\t\t{rf_loss[i]:<10.2f}\t\t{calibrated_rf[i]:<10.2f}\t\t{p_actuals[i]:<10.3f}\n")

            # Create Excel Workbook and Sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Experiment Data"

            # Write the header information
            ws.append(["DEVICE NUMBER", device_num])
            ws.append(["COMMENTS", comment])
            ws.append(["KEITHLEY VOLTAGE", f"{keithley_voltage} V"])
            ws.append(["INITIAL PHOTOCURRENT", f"{initial_current} (mA)"])
            ws.append(["STARTING WAVELENGTH FOR LASER 3", f"{laser_3_WL} (nm)"])
            ws.append(["STARTING WAVELENGTH FOR LASER 4", f"{laser_4_cal} (nm)"])
            ws.append(["DELAY", f"{delay} (s)"])
            ws.append(["DATE", time.strftime("%m/%d/%Y")])
            ws.append(["TIME", time.strftime("%H:%M:%S")])
            ws.append(["RUN TIME", f"{run_time} (s)"])
            ws.append([])  # Add an empty row for spacing

            # Write the table header
            ws.append(["F_BEAT (GHz)", "PHOTOCURRENT (mA)", "Raw RF POW (dBm)", "RF Loss (dB)", "Cal RF POW (dBm)", "VOA P Actual (dBm)"])

            # Write the data rows
            for i in range(len(steps)):
                ws.append([
                    f"{beat_freqs_pow[i]:.2f}",
                    f"{float(photo_currents[i]):.4e}",
                    f"{powers[i]:.2f}",
                    f"{rf_loss[i]:.2f}",
                    f"{calibrated_rf[i]:.2f}",
                    f"{p_actuals[i]:.3f}"
                ])

            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)  # Convert the column to a list
                for cell in column:
                    try:
                        # Compute the length of the cell value
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add a little extra space
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save the workbook
            excel_file_path = file_path.replace(".txt", ".xlsx")
            wb.save(excel_file_path)

            print(f"Data saved to {file_path} and {excel_file_path}")
        else:
            messagebox.showwarning("Cancelled", "Save operation cancelled")

    # Plot the final data

    # Update the plots with the new data
    line1.set_data(steps, beat_freqs)
    line2.set_data(steps, laser_4_wavelengths)
    line3.set_data(beat_freqs, [x[1] for x in beat_freq_and_power])
    line4.set_data(beat_freqs_pow, photo_currents)
    line5.set_data(beat_freqs_pow, calibrated_rf)

    ax1.relim()
    ax2.relim()
    ax3.relim()
    ax4.relim()
    ax5.relim()

    ax1.autoscale_view()
    ax2.autoscale_view()
    ax3.autoscale_view()
    ax4.autoscale_view()
    ax5.autoscale_view()

    # Manually set y-ticks for the RF power graphs with 3 units interval
    rf_power_min = min(powers)
    rf_power_max = max(powers)
    rf_power_ticks = np.arange(np.floor(rf_power_min / 3) * 3, np.ceil(rf_power_max / 3) * 3 + 3, 3)  # Generate y-ticks with 3 units interval

    ax3.set_yticks(rf_power_ticks)
    ax5.set_yticks(rf_power_ticks)

    # Manually set y-ticks for the calibrated RF power graph with 3 units interval
    calibrated_rf_min = min(calibrated_rf)
    calibrated_rf_max = max(calibrated_rf)
    calibrated_rf_ticks = np.arange(np.floor(calibrated_rf_min / 3) * 3, np.ceil(calibrated_rf_max / 3) * 3 + 3, 3)  # Generate y-ticks with 3 units interval

    ax5.set_yticks(calibrated_rf_ticks)

    # Adjust subplot parameters to add space for comments
    plt.subplots_adjust(top=0.85)

    # Add comments to the plot
    comments = [
        f"Device Number: {device_num}",
        f"Comments: {comment}",
        f"Date: {time.strftime('%m/%d/%Y')}",
        f"Time: {time.strftime('%H:%M:%S')}",
        f"Run Time: {run_time:.2f} s",
        f"Keithley Voltage: {keithley_voltage} V",
        f"Excel Loss File: {excel_filename if 'excel_filename' in locals() else 'None'}",
        f"S2P Loss File: {s2p_filename if 's2p_filename' in locals() else 'None'}"
    ]

    # Position for comments (adjust as needed)
    x_comment = 0.5  # X position for the comments
    y_comment_start = 0.94  # Starting Y position for the comments
    y_comment_step = 0.02  # Step size for Y position

    for i, comment in enumerate(comments):
        plt.figtext(x_comment, y_comment_start - i * y_comment_step, comment, wrap=True, horizontalalignment='center', fontsize=10)

    # Maximize the window for different backends
    manager = plt.get_current_fig_manager()
    try:
        # TkAgg backend
        manager.window.state('zoomed')
    except AttributeError:
        try:
            # Qt5Agg backend
            manager.window.showMaximized()
        except AttributeError:
            try:
                # WxAgg backend
                manager.window.Maximize(True)
            except AttributeError:
                pass  # If none of these work, just continue

    # Adjust title and axis label font properties
    title_font = {'size': '14', 'weight': 'bold'}
    label_font = {'size': '12', 'weight': 'bold'}

    ax1.set_title('Beat Frequency vs Step Number', fontdict=title_font)
    ax1.set_xlabel('Step Number', fontdict=label_font)
    ax1.set_ylabel('Beat Frequency (GHz)', fontdict=label_font)
    ax2.set_ylabel('Laser 4 Wavelength (nm)', fontdict=label_font)

    ax3.set_title('Raw RF Power vs Beat Frequency', fontdict=title_font)
    ax3.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
    ax3.set_ylabel('Raw RF Power (dBm)', fontdict=label_font)

    ax4.set_title('Measured Photocurrent vs Beat Frequency', fontdict=title_font)
    ax4.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
    ax4.set_ylabel('Photocurrent (mA)', fontdict=label_font)

    ax5.set_title('Calibrated RF Power vs Beat Frequency', fontdict=title_font)
    ax5.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
    ax5.set_ylabel('Calibrated RF Power (dBm)', fontdict=label_font)

    # Adjust layout
    plt.tight_layout(rect=[0, 0, 1, 0.85])

    plt.draw()
    if output_input in ['YES', 'Y']:
        # Save the plot to the same path as the .txt file but with a .png extension
        plot_file_path = file_path.rsplit('.', 1)[0] + '.png'
        fig.savefig(plot_file_path, bbox_inches='tight')
        print(f"Plot saved to {plot_file_path}")

    # Show the updated plot in the plot window
    plt.show()

    

except KeyboardInterrupt:
    print("Program interrupted by user. Exiting...")
    exit_program()
    
finally:
    # Ensure all resources are closed
    exit_program()
    sys.exit(0)