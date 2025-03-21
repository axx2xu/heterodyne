import time
import math
import pyvisa
import sys
import numpy as np
import openpyxl
from openpyxl import Workbook
import threading
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import mplcursors
from matplotlib.figure import Figure
from matplotlib.ticker import FuncFormatter

# Initialize the VISA resource manager
rm = pyvisa.ResourceManager()

# List all connected VISA devices (Optional: To verify connections)
print("Connected devices:", rm.list_resources())

# Create a VISA adapter for the ECL laser and wavelength meter
# *** The GPIB should always be the same, so these should not need to be changed ***
ecl_adapter_GPIB = 'GPIB0::10::INSTR'  # Update with your actual GPIB address
wavelength_meter_GPIB = 'GPIB0::20::INSTR'  # Update with your actual GPIB address
spectrum_analyzer_GPIB = 'GPIB0::18::INSTR'  # Update with your actual GPIB address
keithley_GPIB = 'GPIB0::24::INSTR'  # Update with your actual GPIB address
RS_power_sensor_GPIB = 'RSNRP::0x00a8::100940::INSTR'  # Update with your actual VISA address for the RS NRP-Z58 sensor
voa_GPIB = 'GPIB0::26::INSTR'  # Update with your actual GPIB address


def exit_program():
    """Exit program and close the connection."""
    ecl_adapter.close()
    wavelength_meter.close()
    spectrum_analyzer.close()
    keithley.close()
    voa.close()
    RS_power_sensor.close()
    update_message_feed("Connection closed.")
    sys.exit(0)


def measure_peak_frequency(spectrum_analyzer):
    """Measure the peak frequency using the spectrum analyzer."""
    try:
        spectrum_analyzer.write('MKPK HI')
        time.sleep(0.1)  # Allow time for the peak search to complete
        peak_freq_1 = spectrum_analyzer.query('MKF?')
        time.sleep(0.1)
        spectrum_analyzer.write('MKPK HI')
        time.sleep(0.1)
        peak_freq_2 = spectrum_analyzer.query('MKF?')
        time.sleep(0.1)
        spectrum_analyzer.write('MKPK HI')
        time.sleep(0.1)
        peak_freq_3 = spectrum_analyzer.query('MKF?')
        peak_freq = min(peak_freq_1, peak_freq_2, peak_freq_3)
        return float(peak_freq) / 1e9  # Convert Hz to GHz
    except Exception as e:
        update_message_feed(f"Error measuring peak frequency with spectrum analyzer: {e}")
        return None


def measure_wavelength_beat(wavelength_meter):
    """Measure the beat frequency using the wavelength meter."""
    try:
        wavelength_meter.write(":INIT:IMM")  # Initiate a single measurement
        time.sleep(1.5)  # Increased wait time for the measurement to complete
        freq_data = wavelength_meter.query(":CALC3:DATA? FREQuency").strip().split(',')
        freqs = [float(freq) for freq in freq_data]
        if len(freqs) == 2:
            return abs(freqs[0] - freqs[1]) / 1e9  # Calculate the difference in GHz
    except Exception as e:
        update_message_feed(f"Error measuring beat frequency with wavelength meter: {e}")
        return None


def set_laser_wavelength(ecl_adapter, channel, wavelength):
    """Set the laser wavelength."""
    update_message_feed(f"Setting laser {channel} wavelength to {wavelength:.3f} nm...")
    ecl_adapter.write(f"CH{channel}:L={wavelength:.3f}")


def read_excel_data(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    frequency = []
    loss = []
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        frequency.append(row[0])
        loss.append(row[1])
    return np.array(frequency), np.array(loss)


def read_s2p_file(filepath):
    frequencies = []
    s12 = []
    s21 = []
    with open(filepath, 'r') as file:
        lines = file.readlines()
    freq_unit = 'hz'
    data_format = 'db'
    for line in lines:
        if line.startswith('#'):
            parts = line.split()
            for part in parts:
                if part.lower() in ['hz', 'khz', 'mhz', 'ghz']:
                    freq_unit = part.lower()
                elif part.lower() in ['ri', 'db', 'ma']:
                    data_format = part.upper()
            continue
        if not line.startswith('!') and not line.startswith('#'):
            values = re.split(r'\s+', line.strip())
            if len(values) >= 9:
                frequencies.append(float(values[0]))
                s21.append(float(values[3]))  # S21 in dB
                s12.append(float(values[5]))  # S12 in dB
    frequencies = np.array(frequencies)
    s_avg = (np.array(s12) + np.array(s21)) / 2
    if freq_unit == 'khz':
        frequencies = frequencies / 1e6
    elif freq_unit == 'mhz':
        frequencies = frequencies / 1e3
    elif freq_unit == 'hz':
        frequencies = frequencies / 1e9
    return frequencies, s_avg


def select_s2p_file():
    file_path = filedialog.askopenfilename(
        title="Select .s2p File",
        filetypes=[("s2p files", "*.s2p"), ("All files", "*.*")]
    )
    if file_path:
        s2p_file_var.set(file_path)


def select_excel_file():
    file_path = filedialog.askopenfilename(
        title="Select .xlsx File",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        excel_file_var.set(file_path)


def custom_linear_interpolation(x, y, x_new):
    """Perform linear interpolation on given data points."""
    x = np.asarray(x)
    y = np.asarray(y)
    x_new = np.asarray(x_new)
    y_new = np.zeros_like(x_new)
    for i, xi in enumerate(x_new):
        if xi <= x[0]:
            y_new[i] = y[0]
        elif xi >= x[-1]:
            y_new[i] = y[-1]
        else:
            k = np.searchsorted(x, xi) - 1
            x_k = x[k]
            x_k1 = x[k + 1]
            y_k = y[k]
            y_k1 = y[k + 1]
            y_new[i] = y_k + (y_k1 - y_k) * (xi - x_k) / (x_k1 - x_k)
    return y_new


def update_message_feed(message):
    """Update the message feed with a new message."""
    message_feed.insert(tk.END, message + "\n")
    message_feed.see(tk.END)  # Scroll to the latest message
    root.update_idletasks()  # Update the GUI to reflect the new message


def calculate_calibrated_rf(powers, beat_freqs_pow, s2p_filename=None, excel_filename=None):
    # Calculate the calibrated RF power by applying RF losses from a network analyzer (.s2p file)
    # and probe losses from an Excel file.
    calibrated_rf = np.array(powers)  # Initialize the calibrated RF power with the raw RF power
    rf_loss = np.zeros_like(calibrated_rf)  # Initialize the list to store the RF loss values
    rf_probe_loss = np.zeros_like(calibrated_rf)  # Initialize the list to store the RF probe loss values
    rf_link_loss = np.zeros_like(calibrated_rf)  # Initialize the list to store the RF link loss values

    # Apply calibration from .s2p file
    if s2p_filename:
        try:
            frequencies, s_avg = read_s2p_file(s2p_filename)
            interpolated_loss = custom_linear_interpolation(frequencies, s_avg, beat_freqs_pow)
            interpolated_loss = np.array(interpolated_loss)
            rf_loss += np.abs(interpolated_loss.real)
            rf_probe_loss += np.abs(interpolated_loss.real)
            calibrated_rf += np.abs(interpolated_loss.real)
            calibrated_rf = np.round(calibrated_rf, 2)
        except Exception as e:
            update_message_feed(f"No network analyzer file found or error in processing: {e}")

    # Apply calibration from Excel file
    if excel_filename:
        try:
            probe_loss_frequency, probe_loss = read_excel_data(excel_filename)
            interpolated_probe_loss = custom_linear_interpolation(probe_loss_frequency, probe_loss, beat_freqs_pow)
            interpolated_probe_loss = np.array(interpolated_probe_loss)
            rf_loss += np.abs(interpolated_probe_loss)
            rf_link_loss += np.abs(interpolated_probe_loss)
            calibrated_rf += np.abs(interpolated_probe_loss)
            calibrated_rf = np.round(calibrated_rf, 2)
        except Exception as e:
            update_message_feed(f"No excel file found or error in processing: {e}")

    return calibrated_rf, rf_loss, rf_probe_loss, rf_link_loss


# Start threading events
data_ready_event = threading.Event()
stop_event = threading.Event()

# Initialize the main Tkinter window
root = tk.Tk()
root.title("Measurement and Plotting GUI")
root.geometry("1200x800")  # Adjust the size as needed
root.state('zoomed')  # Make the window fullscreen

# Create a frame for user inputs on the left side
input_frame = ttk.Frame(root, width=200)
input_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

# Create a frame for the plots on the right side
plot_frame = ttk.Frame(root)
plot_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

# Create input labels and fields
tk.Label(input_frame, text="Starting WL Laser 3 (nm):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
laser_3_var = tk.DoubleVar(value=1550)
laser_3_entry = ttk.Entry(input_frame, textvariable=laser_3_var)
laser_3_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Starting WL Laser 4 (nm):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
laser_4_var = tk.DoubleVar(value=1548)
laser_4_entry = ttk.Entry(input_frame, textvariable=laser_4_var)
laser_4_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Starting Beat Frequency (GHz):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
start_freq_var = tk.DoubleVar(value=0)
start_freq_entry = ttk.Entry(input_frame, textvariable=start_freq_var)
start_freq_entry.grid(row=2, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Ending Beat Frequency (GHz):").grid(row=3, column=0, padx=5, pady=5, sticky="e")
end_freq_var = tk.DoubleVar(value=0)
end_freq_entry = ttk.Entry(input_frame, textvariable=end_freq_var)
end_freq_entry.grid(row=3, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Number of Steps:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
num_steps_var = tk.IntVar(value=0)
num_steps_entry = ttk.Entry(input_frame, textvariable=num_steps_var)
num_steps_entry.grid(row=4, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Delay Between Steps (s):").grid(row=5, column=0, padx=5, pady=5, sticky="e")
delay_var = tk.DoubleVar(value=3.0)
delay_entry = ttk.Entry(input_frame, textvariable=delay_var)
delay_entry.grid(row=5, column=1, padx=5, pady=5)

"""
tk.Label(input_frame, text="Start Frequency Search Threshold (GHz):").grid(row=6, column=0, padx=5, pady=5, sticky="e")
freq_threshold_var = tk.DoubleVar(value=1)
freq_threshold_entry = ttk.Entry(input_frame, textvariable=freq_threshold_var)
freq_threshold_entry.grid(row=6, column=1, padx=5, pady=5)
tk.Label(input_frame, text="Note: Values below 0.5 GHz are less likely to work").grid(row=7, column=0, columnspan=2, pady=2)
"""

# Add the label and button for s2p file selection
tk.Label(input_frame, text="RF Link Loss (.s2p):").grid(row=8, column=0, padx=5, pady=5, sticky="e")
s2p_file_var = tk.StringVar()
s2p_file_entry = ttk.Entry(input_frame, textvariable=s2p_file_var, width=20)
s2p_file_entry.grid(row=8, column=1, padx=5, pady=5)
s2p_file_button = ttk.Button(input_frame, text="Browse", command=select_s2p_file)
s2p_file_button.grid(row=8, column=2, padx=5, pady=5)

# Add the label and button for Excel file selection
tk.Label(input_frame, text="RF Probe Loss (.xlsx):").grid(row=9, column=0, padx=5, pady=5, sticky="e")
excel_file_var = tk.StringVar()
excel_file_entry = ttk.Entry(input_frame, textvariable=excel_file_var, width=20)
excel_file_entry.grid(row=9, column=1, padx=5, pady=5)
excel_file_button = ttk.Button(input_frame, text="Browse", command=select_excel_file)
excel_file_button.grid(row=9, column=2, padx=5, pady=5)

tk.Label(input_frame, text="Enable Automatic Start Beat Frequency Search:").grid(row=10, column=0, padx=5, pady=5, sticky="e")
enable_search_var = tk.BooleanVar(value=True)  # Default is True (enabled)
enable_search_checkbox = ttk.Checkbutton(input_frame, variable=enable_search_var)
enable_search_checkbox.grid(row=10, column=1, padx=5, pady=5, sticky="w")

# Create control buttons below the inputs
start_button = ttk.Button(input_frame, text="START", command=lambda: threading.Thread(target=data_collection).start())
start_button.grid(row=11, column=0, columnspan=2, pady=10)

stop_button = ttk.Button(input_frame, text="STOP", command=lambda: on_stop())
stop_button.grid(row=12, column=0, columnspan=2, pady=10)
tk.Label(input_frame, text="NOTE: This will only stop data collection during the frequency sweep").grid(row=13, column=0, columnspan=2, pady=2)

# Add a Text widget to display live messages
message_feed = tk.Text(input_frame, wrap="word", height=10, width=50)
message_feed.grid(row=14, column=0, columnspan=3, padx=5, pady=5)

cancel_button = ttk.Button(input_frame, text="RESET", command=lambda: on_cancel())
cancel_button.grid(row=15, column=0, columnspan=2, pady=10)
tk.Label(input_frame, text="NOTE: This will reset the program and no data will be saved").grid(row=16, column=0, columnspan=2, pady=2)

# Initialize Matplotlib Figure and Axes
fig = Figure(figsize=(8, 6))
axes = fig.subplots(2, 2)
fig.suptitle("Measurement Plots", fontsize=16)
ax1, ax3, ax4, ax5 = axes[0, 0], axes[0, 1], axes[1, 0], axes[1, 1]
ax2 = ax1.twinx()  # Create a twin y-axis for the first subplot

# Embed the Matplotlib figure into Tkinter
canvas = FigureCanvasTkAgg(fig, master=plot_frame)
canvas.draw()
canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

# Optional: Add Matplotlib Navigation Toolbar
toolbar = NavigationToolbar2Tk(canvas, plot_frame)
toolbar.update()
canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

# Initialize data lists for plotting
steps = []
beat_freqs = []
laser_4_wavelengths = []
beat_freq_and_power = []  # List of tuples: (beat_freq, output_dbm, current, p_actual)
calibrated_rf = []
photo_currents = []
rf_loss = []
rf_probe_loss = []
rf_link_loss = []
powers = []
p_actuals = []

# Setup first subplot (ax1)
color1 = 'tab:blue'
ax1.set_xlabel('Step Number')
ax1.set_ylabel('Beat Frequency (GHz)', color=color1)
ax1.set_title('Beat Frequency vs Step Number')
line1, = ax1.plot([], [], linestyle='-', color=color1)  # Line without markers
markers1, = ax1.plot([], [], 'o', color=color1)  # Markers only
ax1.tick_params(axis='y', labelcolor=color1)
ax1.grid(True)

color2 = 'tab:red'
ax2.set_ylabel('Laser 4 Wavelength (nm)', color=color2)
line2, = ax2.plot([], [], linestyle='--', color=color2)  # Line without markers
markers2, = ax2.plot([], [], 'x', color=color2)  # Markers only
ax2.tick_params(axis='y', labelcolor=color2)
ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x:.3f}'.rstrip('0').rstrip('.')))

# Setup second subplot (ax3)
color3 = 'tab:blue'
ax3.set_xlabel('Beat Frequency (GHz)')
ax3.set_ylabel('Raw RF Power (dBm)', color=color3)
ax3.set_title('Raw RF Power vs Beat Frequency')
line3, = ax3.plot([], [], linestyle='-', color=color3)  # Line without markers
markers3, = ax3.plot([], [], 'o', color=color3)  # Markers only
ax3.tick_params(axis='y', labelcolor=color3)
ax3.grid(True)

# Setup third subplot (ax4)
color4 = 'tab:blue'
ax4.set_xlabel('Beat Frequency (GHz)')
ax4.set_ylabel('Photocurrent (mA)', color=color4)
ax4.set_title('Measured Photocurrent vs Beat Frequency')
line4, = ax4.plot([], [], linestyle='-', color=color4)  # Line without markers
markers4, = ax4.plot([], [], 'o', color=color4)  # Markers only
ax4.tick_params(axis='y', labelcolor=color4)
ax4.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x:.3f}'.rstrip('0').rstrip('.')))
ax4.grid(True)

# Setup fourth subplot (ax5)
color5 = 'tab:blue'
ax5.set_xlabel('Beat Frequency (GHz)')
ax5.set_ylabel('Calibrated RF Power (dBm)', color=color5)
ax5.set_title('Calibrated RF Power vs Beat Frequency')
line5, = ax5.plot([], [], linestyle='-', color=color5)  # Line without markers
markers5, = ax5.plot([], [], 'o', color=color5)  # Markers only
ax5.tick_params(axis='y', labelcolor=color5)
ax5.grid(True)

fig.tight_layout()

# Add hover functionality using mplcursors, but only annotate the markers (actual data points)
mplcursors.cursor(markers1, hover=mplcursors.HoverMode.Transient)
mplcursors.cursor(markers3, hover=mplcursors.HoverMode.Transient)
mplcursors.cursor(markers4, hover=mplcursors.HoverMode.Transient)
mplcursors.cursor(markers5, hover=mplcursors.HoverMode.Transient)


# Create VISA adapters with the connected equipment - send error message to message feed in GUI if connection fails
def open_instruments():
    global ecl_adapter, wavelength_meter, spectrum_analyzer, keithley, RS_power_sensor, voa
    try:
        ecl_adapter = rm.open_resource(ecl_adapter_GPIB)  # Update with your actual GPIB address
        wavelength_meter = rm.open_resource(wavelength_meter_GPIB)  # Update with your actual GPIB address
        spectrum_analyzer = rm.open_resource(spectrum_analyzer_GPIB)  # Update with your actual GPIB address
        keithley = rm.open_resource(keithley_GPIB)  # Update with your actual GPIB address
        keithley.write(":SYSTem:LOCal")
        RS_power_sensor = rm.open_resource(RS_power_sensor_GPIB)  # Update with your actual VISA address for the RS NRP-Z58 sensor
        voa = rm.open_resource(voa_GPIB)  # Update with your actual GPIB address
    except:
        update_message_feed("Error connecting to VISA devices. Check GPIB addresses and connections.")


# Define the data collection function
def data_collection():
    stop_event.clear()
    data_ready_event.clear()

    global steps, beat_freqs, laser_4_wavelengths, beat_freq_and_power, calibrated_rf, photo_currents
    global rf_loss, rf_probe_loss, rf_link_loss, powers, p_actuals
    global start_time, start_time_sweep, sweep_run_time, total_run_time, excel_filename, s2p_filename

    # Create a pop-up window for additional inputs
    input_window = tk.Toplevel(root)
    input_window.title("Save Data Inputs")
    input_window.geometry("300x200")

    # Create input fields for device number and comments
    tk.Label(input_window, text="Device Number:").grid(row=0, column=0, padx=10, pady=10)
    device_num_var = tk.StringVar()
    device_num_entry = ttk.Entry(input_window, textvariable=device_num_var)
    device_num_entry.grid(row=0, column=1, padx=10, pady=10)

    tk.Label(input_window, text="Comments:").grid(row=1, column=0, padx=10, pady=10)
    comment_var = tk.StringVar()
    comment_entry = ttk.Entry(input_window, textvariable=comment_var)
    comment_entry.grid(row=1, column=1, padx=10, pady=10)

    # Function to close the pop-up window and prompt for file path
    def choose_file_and_close():
        global file_path, excel_file_path, plot_file_path, file_path_selected
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if not file_path:
            return  # User cancelled, do not close the window
        excel_file_path = file_path.replace(".txt", ".xlsx")
        plot_file_path = file_path.rsplit('.', 1)[0] + '.png'
        input_window.destroy()

    # Add a button to save the data and close the pop-up
    save_button = ttk.Button(input_window, text="CHOOSE FILE SAVE PATH", command=choose_file_and_close)
    save_button.grid(row=2, column=0, columnspan=2, pady=20)

    try:
        open_instruments()
        time.sleep(5)

        start_time = time.time()

        # Get user inputs for the measurement
        laser_3_WL = laser_3_var.get()
        laser_4_WL = laser_4_var.get()
        num_steps = num_steps_var.get()
        delay = delay_var.get()
        start_freq = start_freq_var.get()
        end_freq = end_freq_var.get()
        enable_search = enable_search_var.get()
        freq_threshold = 0.5  # freq_threshold_var.get()
        excel_filename = excel_file_var.get()
        s2p_filename = s2p_file_var.get()

        # Set the laser wavelengths and power
        set_laser_wavelength(ecl_adapter, 3, laser_3_WL)
        set_laser_wavelength(ecl_adapter, 4, laser_4_WL)

        # Wait for the lasers to stabilize
        update_message_feed("Waiting for the lasers to stabilize...")
        time.sleep(10)

        # Initialize laser wavelengths, frequencies, and other parameters
        c = 299792458  # Speed of light in m/s
        # Set the reference frequency to laser 3
        laser_3_freq = c / (laser_3_WL * 1e-9)  # Convert nm to meters and calculate frequency
        wavelength_meter.write(f":CALC3:DELTA:REF:FREQ {laser_3_freq}")
        time.sleep(1)  # Small delay to ensure command is processed

        # Additional variables to track consecutive increases
        consecutive_increases = 0
        max_consecutive_increases = 3  # Number of consecutive increases to trigger recalibration

        # Measure beat frequency using wavelength meter and ESA
        wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
        esa_beat_freq = measure_peak_frequency(spectrum_analyzer)

        # If wl_meter_beat_freq is None, use ESA value
        if wl_meter_beat_freq is None:
            wl_meter_beat_freq = esa_beat_freq

        current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
        last_beat_freq = None

        if enable_search:
            update_message_feed("RUNNING AUTOMATIC START FREQUENCY SEARCH LOOP...")
            while current_freq >= 0.75:
                if stop_event.is_set():
                    update_message_feed("Data collection stopped by user.")
                    return

                laser_4_freq = c / (laser_4_WL * 1e-9)
                wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
                esa_beat_freq = measure_peak_frequency(spectrum_analyzer)
                if wl_meter_beat_freq is None:
                    wl_meter_beat_freq = esa_beat_freq
                if wl_meter_beat_freq is None or esa_beat_freq is None:
                    continue
                current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq

                if last_beat_freq is not None and current_freq > last_beat_freq:
                    if stop_event.is_set():
                        update_message_feed("Data collection stopped by user.")
                        return
                    if current_freq >= 1:
                        consecutive_increases += 1
                        if consecutive_increases >= max_consecutive_increases:
                            update_message_feed("Start frequency search overshot the threshold, restarting...")
                            laser_4_WL = laser_3_WL - 2  # Reset to initial wavelength - 1 nm
                            set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                            consecutive_increases = 0
                            last_beat_freq = None
                            time.sleep(15)
                            wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
                            esa_beat_freq = measure_peak_frequency(spectrum_analyzer)
                            if wl_meter_beat_freq is None:
                                wl_meter_beat_freq = esa_beat_freq
                            if wl_meter_beat_freq is None or esa_beat_freq is None:
                                continue
                            current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
                            continue
                    elif current_freq < 1:
                        break
                else:
                    consecutive_increases = 0

                if stop_event.is_set():
                    update_message_feed("Data collection stopped by user.")
                    return

                if wl_meter_beat_freq >= 50:
                    update_message_feed(f"Beat Frequency (Wavelength Meter): {wl_meter_beat_freq} GHz")
                    laser_4_new_freq = laser_4_freq - (wl_meter_beat_freq * 0.67 * 1e9)
                    laser_4_WL = (c / laser_4_new_freq) * 1e9
                    if 1540 < laser_4_WL < 1660:
                        set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                    else:
                        update_message_feed(f"New wavelength for laser 4 is out of bounds or unchanged: {laser_4_WL:.3f} nm")
                        exit_program()
                elif esa_beat_freq < 50 and wl_meter_beat_freq < 50:
                    update_message_feed(f"Beat Frequency (ESA): {round(esa_beat_freq,2)} GHz")
                    if esa_beat_freq > 3:
                        if last_beat_freq is not None and last_beat_freq < 1:
                            laser_4_new_freq = laser_4_freq - (0.2 * 1e9)
                        else:
                            laser_4_new_freq = laser_4_freq - (esa_beat_freq * 0.67 * 1e9)
                    elif 1.5 < esa_beat_freq <= 3:
                        laser_4_new_freq = laser_4_freq - (0.5 * 1e9)
                    elif 1 <= esa_beat_freq <= 1.5:
                        laser_4_new_freq = laser_4_freq - (0.2 * 1e9)
                    elif esa_beat_freq < 1:
                        laser_4_new_freq = laser_4_freq - (0.1 * 1e9)
                    laser_4_WL = (c / laser_4_new_freq) * 1e9
                    if 1540 < laser_4_WL < 1660:
                        set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                    else:
                        update_message_feed(f"New wavelength for laser 4 is out of bounds or unchanged: {laser_4_WL:.3f} nm")
                        exit_program()
                last_beat_freq = current_freq
                time.sleep(3)

            if stop_event.is_set():
                update_message_feed("Data collection stopped by user.")
                return

            update_message_feed("Attempting small jump over ESA issues near 0 GHZ...")
            laser_4_freq = c / (laser_4_WL * 1e-9)
            laser_4_new_freq = laser_4_freq - (0.8 * 1e9)
            laser_4_WL = (c / laser_4_new_freq) * 1e9
            set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
            time.sleep(3)
            wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
            esa_beat_freq = measure_peak_frequency(spectrum_analyzer)
            if wl_meter_beat_freq is None:
                wl_meter_beat_freq = esa_beat_freq
            current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq

            if ((not is_valid_number(current_freq)) or current_freq > 2):
                update_message_feed("Attempting second small jump over ESA issues near 0 GHz...")
                laser_4_freq = c / (laser_4_WL * 1e-9)
                laser_4_new_freq = laser_4_freq - (0.4 * 1e9)
                laser_4_WL = (c / laser_4_new_freq) * 1e9
                set_laser_wavelength(4, laser_4_WL)
                time.sleep(3)
                wl_meter_beat_freq = measure_wavelength_beat()
                esa_beat_freq = measure_peak_frequency()
                if wl_meter_beat_freq is None:
                    wl_meter_beat_freq = esa_beat_freq
                current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
                update_message_feed(f"Current Beat Frequency: {round(current_freq,2)} GHz")

            update_laser = True
            while abs(current_freq - start_freq) > freq_threshold:
                if stop_event.is_set():
                    update_message_feed("Data collection stopped by user.")
                    break
                wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
                esa_beat_freq = measure_peak_frequency(spectrum_analyzer)
                if wl_meter_beat_freq is None:
                    wl_meter_beat_freq = esa_beat_freq
                if wl_meter_beat_freq is None or esa_beat_freq is None:
                    continue
                current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
                update_message_feed(f"Current Beat Frequency: {round(current_freq,2)} GHz")
                laser_4_freq = c / (laser_4_WL * 1e-9)
                laser_4_new_freq = laser_4_freq - ((abs(start_freq - current_freq) * 1e9)) / 2
                laser_4_WL = (c / laser_4_new_freq) * 1e9
                if abs(current_freq - start_freq) <= freq_threshold:
                    update_laser = False
                if update_laser:
                    if 1540 < laser_4_WL < 1660:
                        set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                    else:
                        update_message_feed(f"New wavelength for laser 4 is out of bounds or unchanged: {laser_4_WL:.3f} nm")
                        exit_program()
                time.sleep(3)
                last_beat_freq = current_freq

            time.sleep(5)
            wl_meter_beat_freq = measure_wavelength_beat(wavelength_meter)
            esa_beat_freq = measure_peak_frequency(spectrum_analyzer)
            if wl_meter_beat_freq is None:
                wl_meter_beat_freq = esa_beat_freq
            current_freq = wl_meter_beat_freq if wl_meter_beat_freq > 50 else esa_beat_freq
            update_message_feed(f"Current Beat Frequency: {round(current_freq,2)} GHz")
            enable_search = False  # Disable the search after reaching the starting frequency

        # Calculate step size
        laser_4_step = (end_freq - start_freq) / num_steps
        update_message_feed("BEGINNING MEASUREMENT LOOP...")
        start_time_sweep = time.time()
        last_beat_freq = current_freq

        # Get initial photocurrent for text output
        response = keithley.query(":MEASure:CURRent?")
        initial_current_values = response.split(',')
        if len(initial_current_values) > 1:
            initial_current = float(initial_current_values[1]) * 1000  # Convert to mA
            initial_current = round(initial_current, 3)
        keithley.write(":SYSTem:LOCal")

        global looping
        looping = True
        # Begin data collection loop
        for step in range(num_steps):
            if stop_event.is_set():
                update_message_feed("Data collection stopped by user.")
                looping = False
                time_end = time.time()
                sweep_run_time = time_end - start_time_sweep
                total_run_time = time_end - start_time

                beat_freqs, powers, photo_currents, p_actuals = zip(*beat_freq_and_power)
                calibrated_rf, rf_loss, rf_probe_loss, rf_link_loss = calculate_calibrated_rf(
                    powers, beat_freqs, s2p_filename=s2p_filename, excel_filename=excel_filename)
                data_ready_event.set()
                break

            beat_freq = measure_peak_frequency(spectrum_analyzer) if last_beat_freq < 45 else measure_wavelength_beat(wavelength_meter)
            if beat_freq is None:
                continue

            update_message_feed(f"Step {step + 1} of {num_steps}")

            if step < 5 and start_freq < 5 and beat_freq > 10:
                laser_4_freq = c / (laser_4_WL * 1e-9)
                laser_4_new_freq = laser_4_freq - (0.3 * 1e9)
                laser_4_WL = (c / laser_4_new_freq) * 1e9
                set_laser_wavelength(ecl_adapter, 4, laser_4_WL)
                time.sleep(delay)
                beat_freq = measure_peak_frequency(spectrum_analyzer) if last_beat_freq < 45 else measure_wavelength_beat(wavelength_meter)
                if beat_freq is None:
                    continue

            response = keithley.query(":MEASure:CURRent?")
            current_values = response.split(',')
            if len(current_values) > 1:
                current = float(current_values[1]) * 1000  # Convert to mA
                current = round(current, 3)
            keithley.write(":SYSTem:LOCal")

            max_attempts = 3
            attempts = 0
            success = False

            # Measure the VOA P actual
            p_actual = voa.query('READ:POW?')
            p_actual = float(p_actual)
            p_actual = round(p_actual, 3)
            voa.write('SYST:LOC')

            while attempts < max_attempts and not success:
                try:
                    RS_power_sensor.write('INIT:CONT OFF')
                    RS_power_sensor.write('SENS:FUNC "POW:AVG"')
                    RS_power_sensor.write(f'SENS:FREQ {beat_freq}e9')
                    RS_power_sensor.write('SENS:AVER:COUN:AUTO ON')
                    RS_power_sensor.write('SENS:AVER:STAT ON')
                    RS_power_sensor.write('SENS:AVER:TCON REP')
                    RS_power_sensor.write('SENS:POW:AVG:APER 5e-3')

                    rf_outputs = []
                    for i in range(5):
                        RS_power_sensor.write('INIT:IMM')
                        time.sleep(0.1)
                        output = RS_power_sensor.query('TRIG:IMM')
                        output = output.split(',')[0]
                        rf_outputs.append(float(output))
                        time.sleep(0.1)
                    output = sum(rf_outputs) / len(rf_outputs)
                    output_dbm = math.log10(output) * 10 + 30
                    output_dbm = round(output_dbm, 2)
                    beat_freq = round(beat_freq, 2)
                    beat_freq_and_power.append((beat_freq, output_dbm, current, p_actual))
                    success = True
                except ValueError as e:
                    update_message_feed(f"Error processing measurement at step {step + 1}, attempt {attempts + 1}:")
                    attempts += 1
                    time.sleep(1)

            if not success:
                update_message_feed(f"Measurement failed after {max_attempts} attempts at step {step + 1}")
                continue

            update_message_feed(f"Beat Frequency: {round(beat_freq,2)} (GHz)")
            update_message_feed(f"Measured Photocurrent: {current} (mA)")
            update_message_feed(f"Raw RF Power: {output_dbm} (dBm)")

            steps.append(step + 1)
            beat_freqs.append(beat_freq)
            laser_4_wavelengths.append(laser_4_WL)
            rf_loss.append(0)  # Placeholder for RF loss
            calibrated_rf.append(output_dbm)  # Placeholder for calibrated RF
            photo_currents.append(current)
            powers.append(output_dbm)

            laser_4_freq = c / (laser_4_WL * 1e-9)
            laser_4_new_freq = laser_4_freq - (laser_4_step * 1e9)
            laser_4_WL = (c / laser_4_new_freq) * 1e9
            set_laser_wavelength(ecl_adapter, 4, laser_4_WL)

            last_beat_freq = beat_freq
            data_ready_event.set()
            time.sleep(delay)

        looping = False
        update_message_feed("Data collection completed.")
        time_end = time.time()
        sweep_run_time = time_end - start_time_sweep
        total_run_time = time_end - start_time

        beat_freqs, powers, photo_currents, p_actuals = zip(*beat_freq_and_power)
        calibrated_rf, rf_loss, rf_probe_loss, rf_link_loss = calculate_calibrated_rf(
            powers, beat_freqs, s2p_filename=s2p_filename, excel_filename=excel_filename)
        data_ready_event.set()

        device_num = device_num_var.get().strip()
        user_comment = comment_var.get().strip().upper()
        keithley_voltage = keithley.query(":SOUR:VOLT:LEV:IMM:AMPL?").strip()
        keithley_voltage = f"{float(keithley_voltage):.3e}"
        keithley.write(":SYSTem:LOCal")

        fig.subplots_adjust(top=0.8)
        comments = [
            f"Device Number: {device_num}",
            f"Comments: {user_comment}",
            f"Date: {time.strftime('%m/%d/%Y')}",
            f"Time: {time.strftime('%H:%M:%S')}",
            f"Frequency Sweep Run Time: {sweep_run_time:.2f} s",
            f"Total Run Time: {total_run_time:.2f} s",
            f"Keithley Voltage: {keithley_voltage} V",
            f"Excel Loss File: {excel_filename if 'excel_filename' in globals() else 'None'}",
            f"S2P Loss File: {s2p_filename if 's2p_filename' in globals() else 'None'}"
        ]

        x_comment = 0.5
        y_comment_start = 0.94
        y_comment_step = 0.02
        for i, comment in enumerate(comments):
            fig.text(x_comment, y_comment_start - i * y_comment_step, comment, wrap=True,
                     horizontalalignment='center', fontsize=10)

        try:
            root.state('zoomed')
        except AttributeError:
            try:
                root.attributes('-fullscreen', True)
            except AttributeError:
                pass

        title_font = {'size': '14', 'weight': 'bold'}
        label_font = {'size': '12', 'weight': 'bold'}

        ax1.set_title('Beat Frequency vs Step Number', fontdict=title_font)
        ax1.set_xlabel('Step Number', fontdict=label_font)
        ax1.set_ylabel('Beat Frequency (GHz)', fontdict=label_font)
        ax2.set_ylabel('Laser 4 Wavelength (nm)', fontdict=label_font)

        ax3.set_title('Raw RF Power vs Beat Frequency', fontdict=title_font)
        ax3.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
        ax3.set_ylabel('Raw RF Power (dBm)', fontdict=label_font)

        ax4.set_title('Measured Photocurrent vs Beat Frequency', fontdict=title_font)
        ax4.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
        ax4.set_ylabel('Photocurrent (mA)', fontdict=label_font)

        ax5.set_title('Calibrated RF Power vs Beat Frequency', fontdict=title_font)
        ax5.set_xlabel('Beat Frequency (GHz)', fontdict=label_font)
        ax5.set_ylabel('Calibrated RF Power (dBm)', fontdict=label_font)

        min_power = min(powers)
        max_power = max(powers)
        yticks = np.arange(np.floor(min_power / 3) * 3, np.ceil(max_power / 3) * 3 + 3, 3)
        ax3.set_yticks(yticks)
        ax3.set_ylim(min(yticks), max(yticks))
        
        min_calibrated_rf = min(calibrated_rf)
        max_calibrated_rf = max(calibrated_rf)
        yticks = np.arange(np.floor(min_calibrated_rf / 3) * 3, np.ceil(max_calibrated_rf / 3) * 3 + 3, 3)
        ax5.set_yticks(yticks)
        ax5.set_ylim(min(yticks), max(yticks))
        
        fig.tight_layout(rect=[0, 0, 1, 0.85])
        canvas.draw()

        laser_3_WL = laser_3_var.get()
        with open(file_path, 'w') as f:
            f.write("DEVICE NUMBER: " + str(device_num) + "\n")
            f.write("COMMENTS: " + user_comment + "\n")
            f.write("KEITHLEY VOLTAGE: " + str(keithley_voltage) + " V" + "\n")
            f.write("FREQUENCY SWEEP RUN TIME: " + str(f"{sweep_run_time:.2f}") + " s" + "\n")
            f.write("TOTAL RUN TIME: " + str(f"{total_run_time:.2f}") + " s" + "\n")
            f.write("RF Link Loss File (.xlsx): " + str(excel_filename if 'excel_filename' in globals() else 'None') + "\n")
            f.write("RF Probe Loss File (.s2p): " + str(s2p_filename if 's2p_filename' in globals() else 'None') + "\n")
            f.write("INITIAL PHOTOCURRENT: " + str(photo_currents[0]) + " (mA)" + "\n")
            f.write("STARTING WAVELENGTH FOR LASER 3: " + str(laser_3_WL) + " (nm) :" + " STARTING WAVELENGTH FOR LASER 4: " + str(f"{laser_4_wavelengths[0]:.3f}") + " (nm) :" + " DELAY: " + str(delay_var.get()) + " (s) " + "\n")
            f.write("DATE: " + time.strftime("%m/%d/%Y") + "\n")
            f.write("TIME: " + time.strftime("%H:%M:%S") + "\n")
            f.write("\n")
            f.write("F_BEAT(GHz)\tI_PD (mA)\tRaw RF POW (dBm)\tTotal RF Loss (dB)\tProbe RF Loss (dB)\tLink RF Loss (dB)\tCal RF POW (dBm)\tVOA P Actual (dBm)\n")
            for i in range(len(steps)):
                f.write(f"{beat_freqs[i]:<10.2f}\t{photo_currents[i]:<10.4e}\t\t{powers[i]:<10.2f}\t\t{rf_loss[i]:<10.2f}\t\t{rf_probe_loss[i]:<10.2f}\t\t{rf_link_loss[i]:<10.2f}\t\t{calibrated_rf[i]:<10.2f}\t\t{p_actuals[i]:<10.3f}\n")

        fig.savefig(plot_file_path, bbox_inches='tight')
        update_message_feed(f"Data and plot saved to {file_path} and {plot_file_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Data"

        ws.append(["DEVICE NUMBER", device_num])
        ws.append(["COMMENTS", user_comment])
        ws.append(["KEITHLEY VOLTAGE", f"{keithley_voltage} V"])
        ws.append(["INITIAL PHOTOCURRENT", f"{photo_currents[0]} (mA)"])
        ws.append(["STARTING WAVELENGTH FOR LASER 3", f"{laser_3_WL} (nm)"])
        ws.append(["STARTING WAVELENGTH FOR LASER 4", f"{laser_4_wavelengths[0]:.3f} (nm)"])
        ws.append(["DELAY", f"{delay_var.get()} (s)"])
        ws.append(["FREQUENCY SWEEP RUN TIME", f"{sweep_run_time:.2f} s"])
        ws.append(["TOTAL RUN TIME", f"{total_run_time:.2f} s"])
        ws.append(["EXCEL LOSS FILE", excel_filename if 'excel_filename' in globals() else 'None'])
        ws.append(["S2P LOSS FILE", s2p_filename if 's2p_filename' in globals() else 'None'])
        ws.append(["DATE", time.strftime("%m/%d/%Y")])
        ws.append(["TIME", time.strftime("%H:%M:%S")])
        ws.append([])  # Empty row for spacing

        ws.append(["F_BEAT (GHz)", "I_PD (mA)", "Raw RF POW (dBm)", "Total RF Loss (dB)", "RF Probe Loss (dB)", "RF Link Loss (dB)", "Cal RF POW (dBm)", "VOA P Actual (dBm)"])

        for i in range(len(steps)):
            ws.append([
                f"{beat_freqs[i]:.2f}",
                f"{photo_currents[i]}",
                f"{powers[i]:.2f}",
                f"{rf_loss[i]:.2f}",
                f"{rf_probe_loss[i]:.2f}",
                f"{rf_link_loss[i]:.2f}",
                f"{calibrated_rf[i]:.2f}",
                f"{p_actuals[i]:.3f}"
            ])

        for column in ws.columns:
            count = 0
            max_length = 0
            column = list(column)
            for cell in column:
                if count == 1:
                    count += 1
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                count += 1
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
            count += 1

        wb.save(excel_file_path)
        update_message_feed(f"Excel data saved to {excel_file_path}")

    except Exception as e:
        update_message_feed(f"Error in data collection: {e}")
        reset_program()

def is_valid_number(x):
    try:
        # Attempt to convert to a float (this will catch numbers in string form too)
        float(x)
        return True
    except (ValueError, TypeError):
        return False
    
def update_plots():
    if data_ready_event.is_set():
        line1.set_data(steps, beat_freqs)
        markers1.set_data(steps, beat_freqs)
        line2.set_data(steps, laser_4_wavelengths)
        markers2.set_data(steps, laser_4_wavelengths)
        line3.set_data(beat_freqs, powers)
        markers3.set_data(beat_freqs, powers)
        line4.set_data(beat_freqs, photo_currents)
        markers4.set_data(beat_freqs, photo_currents)
        if not looping:
            line5.set_data(beat_freqs, calibrated_rf)
            markers5.set_data(beat_freqs, calibrated_rf)
        ax1.relim()
        ax2.relim()
        ax3.relim()
        ax4.relim()
        ax5.relim()
        ax1.autoscale_view()
        ax2.autoscale_view()
        ax3.autoscale_view()
        ax4.autoscale_view()
        ax5.autoscale_view()
        canvas.draw()
        data_ready_event.clear()

    if not stop_event.is_set():
        root.after(100, update_plots)


def on_stop():
    if messagebox.askyesno("Confirm Stop", "Are you sure you want to stop the data collection?"):
        stop_event.set()
        update_message_feed("Data collection will be stopped.")


def reset_program():
    global steps, beat_freqs, laser_4_wavelengths, beat_freq_and_power, calibrated_rf
    global photo_currents, rf_loss, rf_probe_loss, rf_link_loss, powers, p_actuals, stop_event, data_ready_event
    steps = []
    beat_freqs = []
    laser_4_wavelengths = []
    beat_freq_and_power = []
    calibrated_rf = []
    photo_currents = []
    rf_loss = []
    rf_probe_loss = []
    rf_link_loss = []
    powers = []
    p_actuals = []
    texts_to_remove = [txt for txt in fig.texts if txt != fig._suptitle]
    for txt in texts_to_remove:
        txt.remove()
    line1.set_data([], [])
    markers1.set_data([], [])
    line2.set_data([], [])
    markers2.set_data([], [])
    line3.set_data([], [])
    markers3.set_data([], [])
    line4.set_data([], [])
    markers4.set_data([], [])
    line5.set_data([], [])
    markers5.set_data([], [])
    for ax in [ax1, ax2, ax3, ax4, ax5]:
        ax.relim()
        ax.autoscale_view()
    canvas.draw()
    root.after(100, update_plots)
    stop_event.clear()
    data_ready_event.clear()
    update_message_feed("Program reset and ready to start again.")


def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        stop_event.set()
        root.destroy()
        sys.exit(0)


def on_cancel():
    if messagebox.askyesno("Confirm Exit", "Are you sure you want to reset the program?"):
        stop_event.set()
        time.sleep(4)
        reset_program()


root.protocol("WM_DELETE_WINDOW", on_closing)
root.after(100, update_plots)
root.mainloop()
